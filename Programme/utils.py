import os
import re
from difflib import SequenceMatcher
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

SIGNAL_ROOTS = {
    # Energie & Versorgung (DE & IT)
    "GAS": lambda w: (w.startswith("GAS") and not w.startswith("GAST")) or w.endswith("GAS") or w == "GAS" or "ERDGAS" in w,
    "METANO": lambda w: "METANO" in w,
    "STROM": lambda w: "STROM" in w or "ELETTRIC" in w or "ELETTRO" in w,
    "ENERGIE": lambda w: "ENERGIE" in w or "ENERGIA" in w,
    "WASSER": lambda w: "WASSER" in w or "ACQUA" in w or "IDRIC" in w,
    "FERNWÄRME": lambda w: "FERNWÄRME" in w or "TELERISCALD" in w,
    "PELLETS": lambda w: "PELLET" in w,
    "HOLZ": lambda w: "HOLZ" in w or "LEGNA" in w,
    "MÜLL": lambda w: "MÜLL" in w or "RIFIUT" in w or "TARI" in w,
    # Treibstoffe (DE & IT)
    "DIESEL": lambda w: "DIESEL" in w or "GASOLIO" in w,
    "BENZIN": lambda w: "BENZIN" in w,
    "ADBLUE": lambda w: "ADBLUE" in w,
    "LPG": lambda w: w in ("LPG", "GPL"),
    # Fleisch & Lebensmittel (DE & IT)
    "RIND": lambda w: "RIND" in w or "BOVIN" in w or "MANZO" in w,
    "KALB": lambda w: "KALB" in w or "VITELL" in w,
    "SCHWEIN": lambda w: "SCHWEIN" in w or "SUIN" in w or "MAIALE" in w,
    "GEFLÜGEL": lambda w: "GEFLÜGEL" in w or "POLLAM" in w,
    "HUHN": lambda w: "HUHN" in w or "HÄHNCHEN" in w or "POLLO" in w,
    "LAMM": lambda w: "LAMM" in w or "AGNELL" in w,
    "FISCH": lambda w: "FISCH" in w or "PESCE" in w,
    "BROT": lambda w: "BROT" in w or w == "PANE",
    "MEHL": lambda w: "MEHL" in w or "FARINA" in w,
    "MILCH": lambda w: "MILCH" in w or "LATTE" in w,
    "KÄSE": lambda w: "KÄSE" in w or "FORMAGG" in w,
    # Kostenarten & Dienstleistungen
    "MIETE": lambda w: "MIET" in w or "LOCAZION" in w or "AFFITT" in w,
    "VERSICHERUNG": lambda w: "VERSICHER" in w or "ASSICURAZ" in w,
    "ZINSEN": lambda w: "ZINS" in w or "INTERESS" in w,
    "TELEFON": lambda w: "TELEFON" in w,
    "INTERNET": lambda w: "INTERNET" in w or "CONNECTIV" in w,
    "BERATUNG": lambda w: "BERATUNG" in w or "CONSULENZ" in w,
    "REINIGUNG": lambda w: "REINIG" in w or "PULIZI" in w,
    "WARTUNG": lambda w: "WARTUNG" in w or "MANUTENZ" in w,
    "REPARATUR": lambda w: "REPARATUR" in w or "RIPARAZ" in w,
    "LEASING": lambda w: "LEASING" in w or "NOLEGG" in w,
}

GENERIC_AUXILIARY_KEYWORDS = {
    # Netzausgaben & Transport / Zähler
    "NETZAUSGABEN", "NETZKOSTEN", "NETZENTGELT", "NETZ", "TRASPORTO", "CONTATORE",
    # Systemaufwendungen
    "SYSTEMAUFWENDUNGEN", "SYSTEMKOSTEN", "SYSTEMGEBÜHREN", "ONERI", "ONERE",
    # Steuern, Gebühren & Abgaben
    "STEUERN", "STEUER", "GEBÜHREN", "GEBÜHR", "ABGABEN", "ABGABE", "IMPOSTE", "IMPOSTA", "TASSE", "TASSA", 
    "ACCISE", "ACCISA", "ADDIZIONALE", "CANONE", "BOLLO", "MARCA DA BOLLO",
    # Allgemeine Spesen & Nebenkosten
    "SPESEN", "SPESE", "NEBENKOSTEN", "RUNDUNG", "RUNDUNGEN", "VERSAND", "FRACHT", "PORTO", "SPEDIZIONE", "INCASSO"
}

def extract_signal_words(text: str) -> set:
    if not text:
        return set()
    words = re.findall(r'[A-ZÄÖÜa-zäöü0-9]+', text.upper())
    matched = set()
    for w in words:
        for root, matcher in SIGNAL_ROOTS.items():
            if matcher(w):
                matched.add(root)
    return matched

def is_generic_auxiliary(desc: str) -> bool:
    """Prüft, ob eine Beschreibung ein reiner, unspezifischer Nebenkostenbegriff ist (z.B. 'Steuern', 'Spesen', 'Netzausgaben')."""
    if not desc:
        return False
    desc_clean = desc.upper().strip()
    
    # Wenn ein Träger-Signalwort vorkommt (z.B. GAS, ENERGIE, DIESEL, FLEISCH), ist es KEINE generische Nebenposition
    sig_words = extract_signal_words(desc_clean)
    commodity_signals = sig_words - {"TRANSPORT", "SPEDIZIONE", "FRACHT", "PORTO"}
    if commodity_signals:
        return False
        
    words = re.findall(r'[A-ZÄÖÜa-zäöü0-9]+', desc_clean)
    return any(any(k in w for k in GENERIC_AUXILIARY_KEYWORDS) for w in words)

def is_similar_desc(d1: str, d2: str, threshold: float = 0.80) -> bool:
    if not d1 or not d2:
        return d1 == d2
    d1_clean = d1.strip().upper()
    d2_clean = d2.strip().upper()
    if d1_clean == d2_clean:
        return True
        
    # 1. Hard Veto: Signalwörter prüfen
    sw1 = extract_signal_words(d1_clean)
    sw2 = extract_signal_words(d2_clean)
    if sw1 != sw2:
        return False
        
    # 2. SequenceMatcher Ratio
    ratio = SequenceMatcher(None, d1_clean, d2_clean).ratio()
    if ratio >= threshold:
        return True
        
    # 3. Startswith Check (nur wenn Signalwörter identisch sind und String lang genug)
    if len(d1_clean) > 5 and len(d2_clean) > 5:
        if d1_clean.startswith(d2_clean) or d2_clean.startswith(d1_clean):
            return True
            
    return False

def load_or_create_targa_list(nutzerdaten_dir=None):
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    verkn_ordner = os.path.join(base_dir, "Systemdaten")
        
    os.makedirs(verkn_ordner, exist_ok=True)
        
    targa_file = os.path.join(verkn_ordner, 'TargaListe.xlsx')
    targa_dict = {}
    
    if os.path.exists(targa_file):
        try:
            # Effizienteres Laden mit ExcelFile
            xl = pd.ExcelFile(targa_file)
            sheet_to_read = 'Fahrzeuge' if 'Fahrzeuge' in xl.sheet_names else xl.sheet_names[0]
            df = xl.parse(sheet_to_read)
            
            for index, row in df.iterrows():
                kennzeichen = str(row.get('Kennzeichen', '')).strip().upper().replace(' ', '')
                typ = str(row.get('Typ', '')).strip()
                if kennzeichen and kennzeichen != 'NAN':
                    targa_dict[kennzeichen] = typ
            print(f"Erfolgreich {len(targa_dict)} Fahrzeuge aus TargaListe.xlsx geladen.")
        except Exception as e:
            print(f"Fehler beim Laden der TargaListe.xlsx: {e}")
    else:
        print(f"\nTargaListe.xlsx nicht gefunden. Erstelle Vorlage in {verkn_ordner}...")
        try:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Fahrzeuge"
            ws1['A1'] = "Kennzeichen"
            ws1['B1'] = "Typ"
            ws1.column_dimensions['A'].width = 20
            ws1.column_dimensions['B'].width = 20
            
            ws2 = wb.create_sheet(title="Fahrzeugtypen")
            ws2['A1'] = "Typ"
            default_typen = ["PKW", "LKW", "Transporter", "Traktor", "Motorrad", "Bagger", "Sonstiges"]
            for i, typ in enumerate(default_typen, start=2):
                ws2.cell(row=i, column=1, value=typ)
            ws2.column_dimensions['A'].width = 20
            
            dv = DataValidation(type="list", formula1='=Fahrzeugtypen!$A$2:$A$100', allow_blank=True)
            dv.error = "Bitte wähle einen Typ aus der Liste oder füge ihn im Blatt 'Fahrzeugtypen' hinzu."
            dv.errorTitle = "Ungültiger Typ"
            ws1.add_data_validation(dv)
            dv.add('B2:B1000')
            
            wb.save(targa_file)
            print("=> Vorlage TargaListe.xlsx mit dynamischen Fahrzeugtypen erstellt!\n")
        except Exception as e:
            print(f"Fehler beim Erstellen der TargaListe Vorlage: {e}")
            
    return targa_dict, targa_file

def append_new_targas_to_excel(targa_file, neue_targas_set):
    if not neue_targas_set:
        return
    try:
        wb = load_workbook(targa_file)
        sheet_name = 'Fahrzeuge' if 'Fahrzeuge' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        
        start_row = ws.max_row + 1
        
        for targa in sorted(neue_targas_set):
            ws.cell(row=start_row, column=1, value=targa)
            start_row += 1
            
        wb.save(targa_file)
        print(f"\n=> {len(neue_targas_set)} neue Kennzeichen automatisch zur TargaListe.xlsx hinzugefügt!")
    except Exception as e:
        print(f"\nFehler beim automatischen Speichern der neuen Kennzeichen: {e}")

def ask_shorten_desc():
    try:
        import json
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        config_path = os.path.join(base_dir, "config.json")
        if os.path.exists(config_path):
            with open(config_path, "r", encoding="utf-8") as f:
                config = json.load(f)
                return config.get("shorten_desc", False)
    except Exception as e:
        pass
    return False

def get_text(node, xpath, default=""):
    if node is None:
        return default
    child = node.find(xpath)
    return child.text if child is not None and child.text else default

def safe_float(val, default=0.0, factor=1.0):
    try:
        return float(val) * factor
    except ValueError:
        return default

def read_xml_or_p7m(file_path):
    import io
    if file_path.lower().endswith('.p7m'):
        with open(file_path, 'rb') as f:
            der_data = f.read()
        
        xml_bytes = None
        try:
            from asn1crypto import cms
            content_info = cms.ContentInfo.load(der_data)
            if content_info['content_type'].native == 'signed_data':
                signed_data = content_info['content']
                encap_content_info = signed_data['encap_content_info']
                if encap_content_info['content'].native:
                    xml_bytes = encap_content_info['content'].native
        except Exception:
            pass
            
        if xml_bytes is not None:
            return io.BytesIO(xml_bytes)
            
        # Fallback mit Byte-Slicing
        try:
            start_tags = [b'<?xml', b'<p:FatturaElettronica', b'<FatturaElettronica', b'<ns2:FatturaElettronica', b'<ns3:FatturaElettronica']
            start = -1
            for tag in start_tags:
                start = der_data.find(tag)
                if start != -1:
                    break
                    
            if start != -1:
                end = der_data.rfind(b'>') + 1
                if end > start:
                    return io.BytesIO(der_data[start:end])
        except Exception:
            pass
            
        return None
    else:
        return file_path
