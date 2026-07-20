import sys
import os
import traceback
import re

# 1. Wir versuchen die Module zu laden. Wenn das fehlschlägt, fangen wir den Fehler ab.
import io
try:
    import xml.etree.ElementTree as ET
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font
except ImportError as e:
    print(f"Fehler beim Laden der Module: {e}")
    print("Hast du 'pip install pandas openpyxl' im Terminal ausgeführt?")
    input("\nDrücke Enter zum Beenden...")
    sys.exit(1)

def load_or_create_targa_list():
    script_dir = os.path.dirname(os.path.abspath(__file__)) if '__file__' in globals() else os.getcwd()
    verkn_ordner = os.path.join(script_dir, 'Excel_Verknüpfungen')
    if not os.path.exists(verkn_ordner):
        os.makedirs(verkn_ordner)
        
    targa_file = os.path.join(verkn_ordner, 'TargaListe.xlsx')
    targa_dict = {}
    
    if os.path.exists(targa_file):
        try:
            # Versuche zuerst das 'Fahrzeuge' Blatt zu laden, falls es das neue Format ist.
            # Fallback auf erstes Blatt, falls es das alte Format ist.
            df_dict = pd.read_excel(targa_file, sheet_name=None)
            sheet_to_read = 'Fahrzeuge' if 'Fahrzeuge' in df_dict else list(df_dict.keys())[0]
            df = df_dict[sheet_to_read]
            
            for index, row in df.iterrows():
                kennzeichen = str(row.get('Kennzeichen', '')).strip().upper().replace(' ', '')
                typ = str(row.get('Typ', '')).strip()
                if kennzeichen and kennzeichen != 'NAN':
                    targa_dict[kennzeichen] = typ
            print(f"Erfolgreich {len(targa_dict)} Fahrzeuge aus TargaListe.xlsx geladen.")
        except Exception as e:
            print(f"Fehler beim Laden der TargaListe.xlsx: {e}")
    else:
        print(f"\nTargaListe.xlsx nicht gefunden. Erstelle Vorlage in {verkn_ordner}...")
        try:
            wb = Workbook()
            # Blatt 1: Fahrzeuge
            ws1 = wb.active
            ws1.title = "Fahrzeuge"
            ws1['A1'] = "Kennzeichen"
            ws1['B1'] = "Typ"
            ws1.column_dimensions['A'].width = 20
            ws1.column_dimensions['B'].width = 20
            
            # Blatt 2: Fahrzeugtypen
            ws2 = wb.create_sheet(title="Fahrzeugtypen")
            ws2['A1'] = "Typ"
            default_typen = ["PKW", "LKW", "Transporter", "Traktor", "Motorrad", "Bagger", "Sonstiges"]
            for i, typ in enumerate(default_typen, start=2):
                ws2.cell(row=i, column=1, value=typ)
            ws2.column_dimensions['A'].width = 20
            
            # Data Validation (Dropdown) für Blatt 1, bezieht sich auf Blatt 2
            dv = DataValidation(type="list", formula1='=Fahrzeugtypen!$A$2:$A$100', allow_blank=True)
            dv.error = "Bitte wähle einen Typ aus der Liste oder füge ihn im Blatt 'Fahrzeugtypen' hinzu."
            dv.errorTitle = "Ungültiger Typ"
            ws1.add_data_validation(dv)
            dv.add('B2:B1000')
            
            wb.save(targa_file)
            print("=> Vorlage TargaListe.xlsx mit dynamischen Fahrzeugtypen erstellt!\n")
        except Exception as e:
            print(f"Fehler beim Erstellen der TargaListe Vorlage: {e}")
            
    return targa_dict, targa_file

def append_new_targas_to_excel(targa_file, neue_targas_set):
    if not neue_targas_set:
        return
    try:
        wb = load_workbook(targa_file)
        sheet_name = 'Fahrzeuge' if 'Fahrzeuge' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        
        # Finde die erste leere Zeile in Spalte A
        start_row = 1
        while ws.cell(row=start_row, column=1).value is not None:
            start_row += 1
        
        for targa in sorted(neue_targas_set):
            ws.cell(row=start_row, column=1, value=targa)
            start_row += 1
            
        wb.save(targa_file)
        print(f"\n=> {len(neue_targas_set)} neue Kennzeichen automatisch zur TargaListe.xlsx hinzugefügt!")
    except Exception as e:
        print(f"\nFehler beim automatischen Speichern der neuen Kennzeichen: {e}")

def get_text(node, xpath, default=""):
    if node is None:
        return default
    child = node.find(xpath)
    return child.text if child is not None and child.text else default

def parse_xml_to_list(xml_path, targa_dict, neue_targas_set, fehler_log):
    print(f"Lese: {xml_path}")
    rechnungspositionen = []
    
    try:
        source = xml_path
        if xml_path.lower().endswith('.p7m'):
            with open(xml_path, 'rb') as f:
                data = f.read()
            
            # Suche nach dem Start des XML (verschiedene Tags möglich)
            start_tags = [b'<?xml', b'<p:FatturaElettronica', b'<FatturaElettronica', b'<ns2:FatturaElettronica', b'<ns3:FatturaElettronica']
            start = -1
            for tag in start_tags:
                start = data.find(tag)
                if start != -1:
                    break
            
            if start == -1:
                print(f"Fehler: Konnte keinen XML-Anfang in {xml_path} finden.")
                return []
                
            # Suche nach dem Ende des XML
            end = data.rfind(b'FatturaElettronica>')
            if end == -1:
                print(f"Fehler: Konnte kein XML-Ende in {xml_path} finden.")
                return []
                
            end += len(b'FatturaElettronica>')
            xml_data = data[start:end]
            source = io.BytesIO(xml_data)

        it = ET.iterparse(source)
        for _, el in it:
            if '}' in el.tag:
                el.tag = el.tag.split('}', 1)[1]
        root = it.root

        # 1. Lieferant (CedentePrestatore)
        lieferant_node = root.find('.//CedentePrestatore//DatiAnagrafici//Anagrafica')
        lieferant = "Unbekannter Lieferant"
        if lieferant_node is not None:
            denominazione = get_text(lieferant_node, 'Denominazione')
            if denominazione:
                lieferant = denominazione
            else:
                nome = get_text(lieferant_node, 'Nome')
                cognome = get_text(lieferant_node, 'Cognome')
                if nome or cognome:
                    lieferant = f"{nome} {cognome}".strip()
                    
        liefer_id = get_text(root, './/CedentePrestatore//DatiAnagrafici//IdFiscaleIVA/IdCodice')

        # 2. Rechnungsdaten (DatiGeneraliDocumento)
        dati_generali = root.find('.//DatiGeneraliDocumento')
        rechnungs_datum = get_text(dati_generali, 'Data')
        rechnungs_nummer = get_text(dati_generali, 'Numero')
        waehrung = get_text(dati_generali, 'Divisa', 'EUR')
        
        tipo_documento = get_text(dati_generali, 'TipoDocumento', 'TD01')
        dokumenttyp = "Rechnung"
        faktor = 1.0
        if tipo_documento == 'TD04':
            dokumenttyp = "Gutschrift"
            faktor = -1.0
        elif tipo_documento == 'TD05':
            dokumenttyp = "Belastungsanzeige"
        elif tipo_documento == 'TD08':
            dokumenttyp = "Gutschrift (vereinfacht)"
            faktor = -1.0

        # 3. Kunde (CessionarioCommittente)
        kunde_node = root.find('.//CessionarioCommittente//DatiAnagrafici//Anagrafica')
        kunde = "Unbekannter Kunde"
        if kunde_node is not None:
            k_denominazione = get_text(kunde_node, 'Denominazione')
            if k_denominazione:
                kunde = k_denominazione
            else:
                k_nome = get_text(kunde_node, 'Nome')
                k_cognome = get_text(kunde_node, 'Cognome')
                if k_nome or k_cognome:
                    kunde = f"{k_nome} {k_cognome}".strip()

        kunden_id = get_text(root, './/CessionarioCommittente//DatiAnagrafici//IdFiscaleIVA/IdCodice')

        dati_linee = root.findall('.//DettaglioLinee')
        
        for linea in dati_linee:
            desc = get_text(linea, 'Descrizione', 'Keine Beschreibung')
            desc_short = desc.split(',', 1)[0].strip() if ',' in desc else desc
            qty_text = get_text(linea, 'Quantita', '1.0')
            price_text = get_text(linea, 'PrezzoUnitario', '0.0')
            total_text = get_text(linea, 'PrezzoTotale', '0.0')
            iva_text = get_text(linea, 'AliquotaIVA', '0.0')
            
            # Sichere Konvertierung in Float
            try: qty = float(qty_text) * faktor
            except ValueError: qty = 1.0 * faktor
                
            try: price = float(price_text)
            except ValueError: price = 0.0
                
            try: total = float(total_text) * faktor
            except ValueError: total = 0.0

            try: iva = float(iva_text)
            except ValueError: iva = 0.0
            
            # Extrahiere Targa (Kennzeichen)
            targa_gefunden = ""
            fahrzeugtyp = ""
            altri_dati = linea.findall('.//AltriDatiGestionali')
            for dato in altri_dati:
                tipo = get_text(dato, 'TipoDato').upper()
                if 'TARGA' in tipo or 'TAR' in tipo:
                    targa_gefunden = get_text(dato, 'RiferimentoTesto')
                    break
            if not targa_gefunden and desc:
                # Fallback: Suche in der Beschreibung (z.B. "DIESEL, KZ/Baust.: FL 700 BA - Datum: ...")
                match = re.search(r'KZ/Baust\.:\s*([A-Z0-9\s]+?)\s*(?:-|$)', desc, re.IGNORECASE)
                if match:
                    targa_gefunden = match.group(1).strip()
            
            if targa_gefunden:
                # Normalisiere Kennzeichen für die Suche
                targa_norm = targa_gefunden.strip().upper().replace(' ', '')
                fahrzeugtyp = targa_dict.get(targa_norm, "UNBEKANNT")
                if targa_norm not in targa_dict and targa_norm not in neue_targas_set:
                    neue_targas_set.add(targa_norm)

            abs_path = os.path.abspath(xml_path)
            dateiname = os.path.basename(xml_path)
            hyperlink_formel = f'=HYPERLINK("{abs_path}", "{dateiname}")'

            rechnungspositionen.append({
                'Typ': dokumenttyp,
                'Rechnungsnummer': rechnungs_nummer,
                'Datum': rechnungs_datum,
                'Lieferant': lieferant,
                'Liefer ID': liefer_id,
                'Kunde': kunde,
                'Kunden ID': kunden_id,
                'Beschreibung': desc_short,
                'Kennzeichen': targa_gefunden,
                'Fahrzeugtyp': fahrzeugtyp,
                'Menge': qty,
                f'Einzelpreis ({waehrung})': price,
                f'Gesamtpreis ({waehrung})': total,
                'MwSt (%)': iva / 100.0,
                'Datei': hyperlink_formel
            })
            
    except Exception as e:
        error_msg = f"Fehler beim Parsen von {os.path.basename(xml_path)}: {e}"
        print(error_msg)
        print(traceback.format_exc())
        fehler_log.append(error_msg)
        
    return rechnungspositionen

def run_conversion(paths=None):
    if paths is None:
        paths = sys.argv[1:]
        
    alle_positionen = []
    ausgabe_ordner = ""

    try:
        if len(paths) > 0:
            for pfad in paths:
                # Setze den Ausgabeordner auf das Verzeichnis des ersten Elements
                if not ausgabe_ordner:
                    if os.path.isfile(pfad):
                        ausgabe_ordner = os.path.dirname(pfad)
                    else:
                        ausgabe_ordner = pfad

            # Lade oder erstelle die Targa Liste VOR dem Parsen der XML Dateien
            targa_dict, targa_file = load_or_create_targa_list()
            neue_targas_set = set()
            fehler_log = []

            for pfad in paths:
                if os.path.isfile(pfad) and (pfad.lower().endswith('.xml') or pfad.lower().endswith('.p7m')):
                    alle_positionen.extend(parse_xml_to_list(pfad, targa_dict, neue_targas_set, fehler_log))
                elif os.path.isdir(pfad):
                    print(f"\nDurchsuche Ordner (inkl. Unterordner): {pfad}")
                    for root_dir, _, files in os.walk(pfad):
                        for filename in files:
                            if filename.lower().endswith('.xml') or filename.lower().endswith('.p7m'):
                                alle_positionen.extend(parse_xml_to_list(os.path.join(root_dir, filename), targa_dict, neue_targas_set, fehler_log))
                else:
                    print(f"Überspringe: {pfad} (Keine XML/P7M oder Ordner)")
            
            if alle_positionen:
                print(f"\nErstelle Excel-Datei mit {len(alle_positionen)} Positionen...")
                df = pd.DataFrame(alle_positionen)
                
                # Wenn kein einziges Kennzeichen gefunden wurde, die beiden Spalten entfernen
                has_targa = any(pos.get('Kennzeichen', '') for pos in alle_positionen)
                if not has_targa:
                    if 'Kennzeichen' in df.columns:
                        df = df.drop(columns=['Kennzeichen', 'Fahrzeugtyp'])
                
                # Excel Datei generieren
                script_dir = os.path.dirname(os.path.abspath(__file__)) if '__file__' in globals() else os.getcwd()
                sammlung_ordner = os.path.join(script_dir, 'Excel_Sammlung')
                if not os.path.exists(sammlung_ordner):
                    os.makedirs(sammlung_ordner)
                    
                excel_path = os.path.join(sammlung_ordner, 'Gesammelte_Rechnungen.xlsx')
                
                # Falls die Datei schon existiert, einen eindeutigen Namen finden
                counter = 1
                while os.path.exists(excel_path):
                    excel_path = os.path.join(sammlung_ordner, f'Gesammelte_Rechnungen_{counter}.xlsx')
                    counter += 1

                writer = pd.ExcelWriter(excel_path, engine='openpyxl')
                df.to_excel(writer, index=False, sheet_name='Rechnungen')
                
                worksheet = writer.sheets['Rechnungen']
                # Automatische Spaltenbreite (Perfekte Breite + Puffer)
                for column_cells in worksheet.columns:
                    max_length = 0
                    column_letter = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            if cell.value:
                                val_str = str(cell.value)
                                # Bei Hyperlink-Formeln den angezeigten Text für die Länge verwenden
                                if val_str.startswith('=HYPERLINK'):
                                    parts = val_str.split('", "')
                                    if len(parts) > 1:
                                        val_str = parts[1].replace('")', '').strip()
                                
                                length = len(val_str)
                                if length > max_length:
                                    max_length = length
                        except:
                            pass
                    
                    # Breite = maximale Textlänge + Puffer (ca. 1cm)
                    adjusted_width = max_length + 6 
                    if adjusted_width > 70:  # Spalten nicht unendlich groß machen
                        adjusted_width = 70
                        
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Dynamische Spaltenindizes finden (1-basiert für openpyxl)
                col_indices = {cell.value: idx for idx, cell in enumerate(worksheet[1], start=1)}
                
                einzelpreis_col = next((idx for name, idx in col_indices.items() if name and str(name).startswith('Einzelpreis')), None)
                gesamtpreis_col = next((idx for name, idx in col_indices.items() if name and str(name).startswith('Gesamtpreis')), None)
                mwst_col = col_indices.get('MwSt (%)')
                datei_col = col_indices.get('Datei')

                euro_format = '#,##0.00 €'
                percent_format = '0.00%'
                link_font = Font(color="0563C1", underline="single")
                
                for row in range(2, worksheet.max_row + 1):
                    if einzelpreis_col:
                        worksheet.cell(row=row, column=einzelpreis_col).number_format = euro_format
                    if gesamtpreis_col:
                        worksheet.cell(row=row, column=gesamtpreis_col).number_format = euro_format
                    if mwst_col:
                        worksheet.cell(row=row, column=mwst_col).number_format = percent_format
                    if datei_col:
                        worksheet.cell(row=row, column=datei_col).font = link_font
                
                writer.close()
                
                print(f"Erfolgreich gespeichert unter: {excel_path}")
                
                append_new_targas_to_excel(targa_file, neue_targas_set)
                
                if fehler_log:
                    log_pfad = os.path.join(sammlung_ordner, 'Fehlgeschlagen.txt')
                    with open(log_pfad, 'w', encoding='utf-8') as f:
                        f.write("Folgende Fehler traten beim Verarbeiten auf:\n\n")
                        for err in fehler_log:
                            f.write(f"- {err}\n")
                    print(f"\nAchtung: Es gab Fehler. Details siehe: {log_pfad}")
                
                if os.name == 'nt':
                    os.startfile(excel_path)
                elif sys.platform == 'darwin':
                    os.system(f'open "{excel_path}"')
                else:
                    os.system(f'xdg-open "{excel_path}"')
            else:
                print("\nEs wurden keine gültigen Rechnungspositionen gefunden.")
        else:
            print("Ziehe eine oder mehrere XML- oder P7M-Dateien (Drag & Drop) auf dieses Skript-Icon, um sie zu konvertieren.")
            print("Oder ziehe einen ganzen Ordner mit XML/P7M-Dateien auf das Icon.")
    except Exception as e:
        print("\nEin unerwarteter Fehler ist aufgetreten:")
        print(traceback.format_exc())
        
        # Dieser Befehl hält das Fenster ganz am Schluss offen, egal was passiert ist
        if paths is None or paths == sys.argv[1:]:
            input("\nDrücke Enter zum Beenden...")

if __name__ == "__main__":
    run_conversion()