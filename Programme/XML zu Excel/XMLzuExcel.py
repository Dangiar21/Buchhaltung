import sys
import os
import traceback
import re

# Utils aus dem übergeordneten Ordner laden
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from utils import load_or_create_targa_list, append_new_targas_to_excel, ask_shorten_desc, get_text, safe_float, read_xml_or_p7m

try:
    import defusedxml.ElementTree as ET
    import pandas as pd
except Exception as e:
    print(f"Fehler beim Laden der Module: {e}")
    print("Hast du 'pip install pandas openpyxl defusedxml' im Terminal ausgeführt?")
    input("\nDrücke Enter zum Beenden...")
    sys.exit(1)

def ask_shorten_desc_local_fallback():
    # Wird nun über utils.py importiert.
    pass

from sdi_parser import parse_sdi_xml

def parse_xml_to_list(xml_path, targa_dict=None, neue_targas_set=None, fehler_log=None, shorten_description=True):
    if targa_dict is None: targa_dict = {}
    if neue_targas_set is None: neue_targas_set = set()
    if fehler_log is None: fehler_log = []
    
    print(f"Lese: {xml_path}")
    
    try:
        parsed_items = parse_sdi_xml(xml_path, targa_dict, neue_targas_set, fehler_log, shorten_description)
        rechnungspositionen = []
        
        for item in parsed_items:
            waehrung = item.get('Waehrung', 'EUR')
            
            rechnungspositionen.append({
                'Typ': item['Typ'],
                'Rechnungsnummer': item['Rechnungsnummer'],
                'Datum': item['Datum'],
                'Lieferant': item['Lieferant'],
                'Liefer ID': item['Liefer ID'],
                'Kunde': item['Kunde'],
                'Kunden ID': item['Kunden ID'],
                'Beschreibung': item['Beschreibung'],
                'CdC': item['CdC'],
                'Kennzeichen': item['Kennzeichen'],
                'Fahrzeugtyp': item['Fahrzeugtyp'],
                'Menge': item['Menge'],
                f'Einzelpreis ({waehrung})': item['Einzelpreis_Roh'],
                f'Gesamtpreis ({waehrung})': item['Gesamtpreis_Roh'],
                'MwSt (%)': item['MwSt'],
                'Datei': item['Datei_Link']
            })
            
        return rechnungspositionen

    except Exception as e:
        error_msg = f"Fehler beim Parsen von {os.path.basename(xml_path)}: {e}"
        print(error_msg)
        print(traceback.format_exc())
        fehler_log.append(error_msg)
        return []

def run_conversion(paths=None, output_dir=None, nutzerdaten_dir=None):
    if paths is None:
        paths = sys.argv[1:]
        
    alle_positionen = []
    ausgabe_ordner = output_dir

    try:
        if len(paths) > 0:
            for pfad in paths:
                if not ausgabe_ordner:
                    if os.path.isfile(pfad):
                        ausgabe_ordner = os.path.dirname(pfad)
                    else:
                        ausgabe_ordner = pfad

            targa_dict, targa_file = load_or_create_targa_list(nutzerdaten_dir)
            neue_targas_set = set()
            fehler_log = []
            
            shorten_description = ask_shorten_desc()

            for pfad in paths:
                if os.path.isfile(pfad) and (pfad.lower().endswith('.xml') or pfad.lower().endswith('.p7m')):
                    alle_positionen.extend(parse_xml_to_list(pfad, targa_dict, neue_targas_set, fehler_log, shorten_description))
                elif os.path.isdir(pfad):
                    print(f"\nDurchsuche Ordner (inkl. Unterordner): {pfad}")
                    for root_dir, _, files in os.walk(pfad):
                        for filename in files:
                            if filename.lower().endswith('.xml') or filename.lower().endswith('.p7m'):
                                alle_positionen.extend(parse_xml_to_list(os.path.join(root_dir, filename), targa_dict, neue_targas_set, fehler_log, shorten_description))
                else:
                    print(f"Überspringe: {pfad} (Keine XML/P7M oder Ordner)")
            
            if alle_positionen:
                print(f"\nErstelle Excel-Datei mit {len(alle_positionen)} Positionen...")
                df = pd.DataFrame(alle_positionen)
                
                has_targa = any(pos.get('Kennzeichen', '') for pos in alle_positionen)
                if not has_targa:
                    if 'Kennzeichen' in df.columns:
                        df = df.drop(columns=['Kennzeichen', 'Fahrzeugtyp'])
                
                if output_dir:
                    sammlung_ordner = output_dir
                else:
                    script_dir = os.path.dirname(os.path.abspath(__file__)) if '__file__' in globals() else os.getcwd()
                    sammlung_ordner = os.path.join(script_dir, 'Excel_Sammlung')
                    
                if not os.path.exists(sammlung_ordner):
                    os.makedirs(sammlung_ordner)
                    
                excel_path = os.path.join(sammlung_ordner, 'Gesammelte_XML_Daten.xlsx')
                
                counter = 1
                while os.path.exists(excel_path):
                    excel_path = os.path.join(sammlung_ordner, f'Gesammelte_XML_Daten_{counter}.xlsx')
                    counter += 1

                writer = pd.ExcelWriter(excel_path, engine='openpyxl')
                df.to_excel(writer, index=False, sheet_name='XML_Daten')
                
                worksheet = writer.sheets['XML_Daten']
                for column_cells in worksheet.columns:
                    max_length = 0
                    column_letter = column_cells[0].column_letter
                    for cell in column_cells[:50]:
                        try:
                            if cell.value:
                                val_str = str(cell.value)
                                if val_str.startswith('=HYPERLINK'):
                                    parts = val_str.split('", "')
                                    if len(parts) > 1:
                                        val_str = parts[1].replace('")', '').strip()
                                
                                length = len(val_str)
                                if length > max_length:
                                    max_length = length
                        except Exception as e:
                            print(f'Fehler: {e}')
                            pass
                    
                    adjusted_width = max_length + 6 
                    if adjusted_width > 70:
                        adjusted_width = 70
                        
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                col_indices = {cell.value: idx for idx, cell in enumerate(worksheet[1], start=1)}
                
                einzelpreis_col = next((idx for name, idx in col_indices.items() if name and str(name).startswith('Einzelpreis')), None)
                gesamtpreis_col = next((idx for name, idx in col_indices.items() if name and str(name).startswith('Gesamtpreis')), None)
                mwst_col = col_indices.get('MwSt (%)')
                datei_col = col_indices.get('Datei')

                euro_format = '#,##0.00 €'
                percent_format = '0.00%'
                from openpyxl.styles import Font
                link_font = Font(color="0563C1", underline="single")
                
                for row in range(2, worksheet.max_row + 1):
                    if einzelpreis_col:
                        worksheet.cell(row=row, column=einzelpreis_col).number_format = euro_format
                    if gesamtpreis_col:
                        worksheet.cell(row=row, column=gesamtpreis_col).number_format = euro_format
                    if mwst_col:
                        worksheet.cell(row=row, column=mwst_col).number_format = percent_format
                    if datei_col:
                        worksheet.cell(row=row, column=datei_col).font = link_font
                
                writer.close()
                
                print(f"\n✅ Erfolgreich gespeichert unter: {excel_path}")
                
                append_new_targas_to_excel(targa_file, neue_targas_set)
                
                if fehler_log:
                    log_pfad = os.path.join(sammlung_ordner, 'Fehlgeschlagen.txt')
                    with open(log_pfad, 'w', encoding='utf-8') as f:
                        f.write("Folgende Fehler traten beim Verarbeiten auf:\n\n")
                        for err in fehler_log:
                            f.write(f"- {err}\n")
                    print(f"\nAchtung: Es gab Fehler. Details siehe: {log_pfad}")
                
                if os.name == 'nt' or sys.platform == 'win32':
                    os.startfile(excel_path)
                elif sys.platform == 'darwin':
                    import subprocess
                    subprocess.run(['open', excel_path], check=True)
                else:
                    import subprocess
                    subprocess.run(['xdg-open', excel_path], check=True)
            else:
                print("\nEs wurden keine gültigen Rechnungspositionen gefunden.")
        else:
            print("Ziehe eine oder mehrere XML- oder P7M-Dateien (Drag & Drop) auf dieses Skript-Icon, um sie zu konvertieren.")
            print("Oder ziehe einen ganzen Ordner mit XML/P7M-Dateien auf das Icon.")
    except Exception as e:
        print("\nEin unerwarteter Fehler ist aufgetreten:")
        print(traceback.format_exc())
        
        if paths is None or paths == sys.argv[1:]:
            if sys.stdout.isatty():
                input("\nDrücke Enter zum Beenden...")

if __name__ == "__main__":
    run_conversion()
