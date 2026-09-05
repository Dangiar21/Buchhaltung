import os
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
import pandas as pd
import sys

script_dir = os.path.dirname(os.path.abspath(__file__))
prog_dir = os.path.dirname(script_dir)
if prog_dir not in sys.path:
    sys.path.append(prog_dir)

def normalize_id(val):
    """Bulletproof ID normalizer: handles NaN, float casts (.0), casing, and leading zeros."""
    if pd.isna(val): return ""
    s = str(val).strip().lower()
    if s in ('nan', ''): return ""
    if s.endswith('.0'): s = s[:-2]
    stripped = s.lstrip('0')
    return '0' if not stripped else stripped

def ensure_rule_file(file_path):
    """
    Erstellt oder aktualisiert die Excel-Datei mit Kontenregeln.
    Verwendet einen einheitlichen Reiter 'Regeln' mit den Spalten:
      - Lieferant (Name oder MwSt-Nr)
      - Stichwort in Beschreibung
      - Konto
    Migriert Altdaten aus 'Lieferanten-Regeln' und 'Stichwort-Regeln' automatisch
    und entfernt veraltete Reiter.
    """
    modified = False
    
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        modified = True

    # 1. Einheitlichen Reiter 'Regeln' sicherstellen & ggf. Altdaten migrieren
    if "Regeln" not in wb.sheetnames:
        ws_regeln = wb.create_sheet("Regeln")
        ws_regeln.append(["Lieferant (Name oder MwSt-Nr)", "Stichwort in Beschreibung", "Konto"])
        for cell in ws_regeln[1]:
            cell.font = Font(bold=True)
        ws_regeln.column_dimensions['A'].width = 35
        ws_regeln.column_dimensions['B'].width = 35
        ws_regeln.column_dimensions['C'].width = 15
        modified = True

        # Altdaten aus Lieferanten-Regeln migrieren
        if "Lieferanten-Regeln" in wb.sheetnames:
            ws_old_lief = wb["Lieferanten-Regeln"]
            for row in ws_old_lief.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2:
                    lief = str(row[0]).strip() if row[0] is not None else ""
                    konto = str(row[1]).strip() if row[1] is not None else ""
                    if lief and lief.lower() != 'nan' and konto and konto.lower() != 'nan':
                        ws_regeln.append([lief, "", konto])
                        modified = True

        # Altdaten aus Stichwort-Regeln migrieren
        if "Stichwort-Regeln" in wb.sheetnames:
            ws_old_stich = wb["Stichwort-Regeln"]
            for row in ws_old_stich.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2:
                    stich = str(row[0]).strip() if row[0] is not None else ""
                    konto = str(row[1]).strip() if row[1] is not None else ""
                    if stich and stich.lower() != 'nan' and konto and konto.lower() != 'nan':
                        ws_regeln.append(["", stich, konto])
                        modified = True
    else:
        # Falls 'Regeln' bereits existiert, Spaltenüberschriften prüfen
        ws_regeln = wb["Regeln"]
        if ws_regeln.max_row == 0 or ws_regeln.cell(1, 1).value is None:
            ws_regeln.append(["Lieferant (Name oder MwSt-Nr)", "Stichwort in Beschreibung", "Konto"])
            for cell in ws_regeln[1]:
                cell.font = Font(bold=True)
            ws_regeln.column_dimensions['A'].width = 35
            ws_regeln.column_dimensions['B'].width = 35
            ws_regeln.column_dimensions['C'].width = 15
            modified = True

    # 2. Veraltete Reiter entfernen
    obsolete_sheets = ["Lieferanten-Regeln", "Stichwort-Regeln", "Kontenplan", "Sheet1", "KI-Zuweisungen", "Sheet"]
    for old_sheet in obsolete_sheets:
        if old_sheet in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb[old_sheet]
            modified = True

    # 3. Veraltete DataValidations entfernen
    if "Regeln" in wb.sheetnames:
        ws = wb["Regeln"]
        valid_dvs = []
        for dv in ws.data_validations.dataValidation:
            if dv.formula1 and "Kontenplan" in str(dv.formula1):
                modified = True
            else:
                valid_dvs.append(dv)
        ws.data_validations.dataValidation = valid_dvs

    if modified:
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        try:
            wb.save(file_path)
        except (PermissionError, OSError) as e:
            print(f"Warnung: Konnte {file_path} nicht aktualisieren (evtl. in Excel geöffnet): {e}")

def parse_excel_rules(file_path):
    """
    Parst Regeln aus einer Excel-Datei.
    Unterstützt sowohl den neuen Reiter 'Regeln' (Lieferant, Stichwort, Konto)
    als auch alte 2-Reiter-Dateien als Fallback.
    Gibt eine Liste von Dicts mit {'lieferant', 'suchbegriff', 'konto', 'typ'} zurück.
    """
    rules_list = []
    if not os.path.exists(file_path):
        return rules_list

    try:
        xl = pd.ExcelFile(file_path)
        sheet_to_use = None
        for name in ["Regeln", "Kontenregeln"]:
            if name in xl.sheet_names:
                sheet_to_use = name
                break

        if sheet_to_use:
            df = xl.parse(sheet_to_use)
            lief_col = None
            stich_col = None
            konto_col = None

            for col in df.columns:
                c_lower = str(col).lower()
                if "lieferant" in c_lower or "mwst" in c_lower:
                    lief_col = col
                elif "stichwort" in c_lower or "wort" in c_lower or "beschreib" in c_lower:
                    stich_col = col
                elif "konto" in c_lower:
                    konto_col = col

            for _, row in df.iterrows():
                lief = str(row[lief_col]).strip() if lief_col and pd.notna(row[lief_col]) else ""
                if not lief_col and len(row) > 0 and pd.notna(row.iloc[0]):
                    lief = str(row.iloc[0]).strip()
                if lief.lower() in ('nan', 'none', ''):
                    lief = ""

                stich = str(row[stich_col]).strip() if stich_col and pd.notna(row[stich_col]) else ""
                if not stich_col and len(row) > 1 and pd.notna(row.iloc[1]):
                    stich = str(row.iloc[1]).strip()
                if stich.lower() in ('nan', 'none', ''):
                    stich = ""

                konto_raw = str(row[konto_col]).strip() if konto_col and pd.notna(row[konto_col]) else ""
                if not konto_col and len(row) > 2 and pd.notna(row.iloc[2]):
                    konto_raw = str(row.iloc[2]).strip()

                konto = konto_raw.split(' - ')[0].strip() if konto_raw and konto_raw.lower() not in ('nan', 'none') else ""
                if konto.endswith('.0'):
                    konto = konto[:-2]

                if not konto or (not lief and not stich):
                    continue

                if lief and stich:
                    r_type = "kombi"
                elif lief:
                    r_type = "lieferant"
                else:
                    r_type = "stichwort"

                rules_list.append({
                    "lieferant": lief.lower(),
                    "suchbegriff": stich.lower(),
                    "konto": konto,
                    "typ": r_type
                })
        else:
            # Fallback für ältere Excel-Dateien
            if "Lieferanten-Regeln" in xl.sheet_names:
                df_l = xl.parse("Lieferanten-Regeln")
                for _, row in df_l.iterrows():
                    lief = str(row.iloc[0]).strip().lower() if len(row) > 0 and pd.notna(row.iloc[0]) else ""
                    k_raw = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
                    k = k_raw.split(' - ')[0].strip() if k_raw and k_raw.lower() not in ('nan', 'none') else ""
                    if k.endswith('.0'): k = k[:-2]
                    if lief and lief not in ('nan', '') and k:
                        rules_list.append({"lieferant": lief, "suchbegriff": "", "konto": k, "typ": "lieferant"})

            if "Stichwort-Regeln" in xl.sheet_names:
                df_s = xl.parse("Stichwort-Regeln")
                for _, row in df_s.iterrows():
                    stich = str(row.iloc[0]).strip().lower() if len(row) > 0 and pd.notna(row.iloc[0]) else ""
                    k_raw = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
                    k = k_raw.split(' - ')[0].strip() if k_raw and k_raw.lower() not in ('nan', 'none') else ""
                    if k.endswith('.0'): k = k[:-2]
                    if stich and stich not in ('nan', '') and k:
                        rules_list.append({"lieferant": "", "suchbegriff": stich, "konto": k, "typ": "stichwort"})

    except Exception as e:
        print(f"Fehler beim Parsen der Regeln aus {file_path}: {e}")

    return rules_list

def load_rules(global_path, client_path):
    """
    Lädt die Regeln über den DatabaseManager (CQRS). Excel dient als Frontend.
    Gibt ein strukturiertes Rules-Dictionary mit sortierten Listen zurück:
      - client_kombi (Prio 1)
      - client_lieferant (Prio 2)
      - client_stichwort (Prio 3)
      - global_kombi (Prio 4)
      - global_stichwort (Prio 5)
      - global_lieferant (Prio 6)
    """
    rules = {
        "client_kombi": [],
        "client_lieferant": [],
        "client_stichwort": [],
        "global_kombi": [],
        "global_stichwort": [],
        "global_lieferant": []
    }

    try:
        from DatabaseManager import get_db
        db = get_db()
    except ImportError:
        print("Fehler: DatabaseManager nicht gefunden.")
        return rules

    # 1. Global Sync
    if os.path.exists(global_path):
        global_mtime = os.path.getmtime(global_path)
        last_sync = db.get_sync_status("GLOBAL", "global_rules")

        if global_mtime > last_sync:
            print("Synchronisiere globale Regeln aus Excel in die SQLite-Datenbank...")
            try:
                parsed_global = parse_excel_rules(global_path)
                rules_list = []
                for r in parsed_global:
                    typ = r["typ"]
                    if typ == "kombi":
                        prio = 4
                        reg_typ = "global_kombi"
                    elif typ == "stichwort":
                        prio = 5
                        reg_typ = "global_stichwort"
                    else:
                        prio = 6
                        reg_typ = "global_lieferant"

                    rules_list.append({
                        "prioritaet": prio,
                        "lieferant": r["lieferant"],
                        "suchbegriff": r["suchbegriff"],
                        "konto": r["konto"],
                        "regel_typ": reg_typ
                    })

                df_sync = pd.DataFrame(rules_list)
                db.sync_rules("GLOBAL", "global_rules", df_sync)
                db.set_sync_status("GLOBAL", "global_rules", global_mtime)
            except Exception as e:
                print(f"Fehler beim Sync der globalen Regeln: {e}")

    # 2. Client Sync
    client_id = ""
    if client_path and os.path.exists(client_path):
        client_id = os.path.basename(os.path.dirname(client_path))
        if client_id == "Nutzerdaten":
            client_id = os.path.basename(os.path.dirname(os.path.dirname(client_path)))

        client_mtime = os.path.getmtime(client_path)
        last_sync = db.get_sync_status(client_id, "client_rules")

        needs_sync = client_mtime > last_sync
        if not needs_sync:
            try:
                df_check = db.get_rules(client_id, "client_rules")
                if df_check.empty:
                    needs_sync = True
            except Exception:
                needs_sync = True

        if needs_sync:
            print(f"Synchronisiere kunden-spezifische Regeln für {client_id} aus Excel in SQLite...")
            try:
                parsed_client = parse_excel_rules(client_path)
                rules_list = []
                for r in parsed_client:
                    typ = r["typ"]
                    if typ == "kombi":
                        prio = 1
                        reg_typ = "client_kombi"
                    elif typ == "lieferant":
                        prio = 2
                        reg_typ = "client_lieferant"
                    else:
                        prio = 3
                        reg_typ = "client_stichwort"

                    rules_list.append({
                        "prioritaet": prio,
                        "lieferant": r["lieferant"],
                        "suchbegriff": r["suchbegriff"],
                        "konto": r["konto"],
                        "regel_typ": reg_typ
                    })

                df_sync = pd.DataFrame(rules_list)
                db.sync_rules(client_id, "client_rules", df_sync)
                db.set_sync_status(client_id, "client_rules", client_mtime)
            except Exception as e:
                print(f"Fehler beim Sync der kunden-spezifischen Regeln: {e}")

    # 3. Aus SQLite laden
    try:
        # Global laden
        df_global = db.get_rules("GLOBAL", "global_rules")
        if not df_global.empty:
            for _, row in df_global.iterrows():
                typ = row['regel_typ']
                lief = str(row.get('lieferant', '')).strip().lower() if pd.notna(row.get('lieferant')) else ""
                stich = str(row.get('suchbegriff', '')).strip().lower() if pd.notna(row.get('suchbegriff')) else ""
                konto = str(row['konto']).strip()
                if typ in rules:
                    rules[typ].append({"lieferant": lief, "suchbegriff": stich, "konto": konto})

        # Client laden
        if client_id:
            df_client = db.get_rules(client_id, "client_rules")
            if not df_client.empty:
                for _, row in df_client.iterrows():
                    typ = row['regel_typ']
                    lief = str(row.get('lieferant', '')).strip().lower() if pd.notna(row.get('lieferant')) else ""
                    stich = str(row.get('suchbegriff', '')).strip().lower() if pd.notna(row.get('suchbegriff')) else ""
                    konto = str(row['konto']).strip()
                    if typ in rules:
                        rules[typ].append({"lieferant": lief, "suchbegriff": stich, "konto": konto})

    except Exception as e:
        print(f"Fehler beim Laden der Regeln aus SQLite: {e}")

    return rules

def _match_supplier(rule_lief, supplier_name, raw_supplier_vat, norm_supplier_vat):
    """Prüft, ob ein Lieferant nach Name oder MwSt-Nummer matcht."""
    if not rule_lief:
        return False
    norm_rule = normalize_id(rule_lief)
    return (
        (rule_lief in supplier_name) or
        (raw_supplier_vat and rule_lief in raw_supplier_vat) or
        (norm_rule and norm_rule == norm_supplier_vat)
    )

def _match_keyword(rule_stich, desc, desc_norm):
    """Prüft, ob ein Stichwort in der Beschreibung vorkommt."""
    if not rule_stich:
        return False
    return (rule_stich in desc) or (rule_stich in desc_norm)

def assign_account(desc_norm, desc, supplier_name, supplier_vat, kunden_id, rules):
    """
    Weist das Konto basierend auf der 6-stufigen Priorität zu:
      1. Kunde – Kombination (Lieferant UND Stichwort)
      2. Kunde – Nur Lieferant
      3. Kunde – Nur Stichwort
      4. Global – Kombination (Lieferant UND Stichwort)
      5. Global – Nur Stichwort
      6. Global – Nur Lieferant
    Gibt (Konto, is_pending) zurück.
    """
    desc = str(desc).lower()
    desc_norm = str(desc_norm).lower()
    supplier_name = str(supplier_name).lower()
    norm_supplier_vat = normalize_id(supplier_vat)
    raw_supplier_vat = str(supplier_vat).strip().lower() if supplier_vat else ""

    def _eval_items(item_list, is_kombi, is_lief, is_stich):
        for r in item_list:
            if isinstance(r, dict):
                r_lief = r.get("lieferant", "")
                r_stich = r.get("suchbegriff", "")
                r_konto = r.get("konto", "")
            else:
                continue

            if is_kombi:
                if _match_supplier(r_lief, supplier_name, raw_supplier_vat, norm_supplier_vat) and \
                   _match_keyword(r_stich, desc, desc_norm):
                    return str(r_konto)
            elif is_lief:
                if _match_supplier(r_lief, supplier_name, raw_supplier_vat, norm_supplier_vat):
                    return str(r_konto)
            elif is_stich:
                if _match_keyword(r_stich, desc, desc_norm):
                    return str(r_konto)
        return None

    # 1. Höchste Priorität: Kunde – Kombination
    kto = _eval_items(rules.get("client_kombi", []), is_kombi=True, is_lief=False, is_stich=False)
    if kto: return kto, False

    # 2. Priorität: Kunde – Nur Lieferant
    kto = _eval_items(rules.get("client_lieferant", []), is_kombi=False, is_lief=True, is_stich=False)
    if kto: return kto, False

    # 3. Priorität: Kunde – Nur Stichwort
    kto = _eval_items(rules.get("client_stichwort", []), is_kombi=False, is_lief=False, is_stich=True)
    if kto: return kto, False

    # 4. Priorität: Global – Kombination
    kto = _eval_items(rules.get("global_kombi", []), is_kombi=True, is_lief=False, is_stich=False)
    if kto: return kto, False

    # 5. Priorität: Global – Nur Stichwort
    kto = _eval_items(rules.get("global_stichwort", []), is_kombi=False, is_lief=False, is_stich=True)
    if kto: return kto, False

    # 6. Priorität: Global – Nur Lieferant
    kto = _eval_items(rules.get("global_lieferant", []), is_kombi=False, is_lief=True, is_stich=False)
    if kto: return kto, False

    return "???", False
