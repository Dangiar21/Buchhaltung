import pandas as pd
import os
from openpyxl.styles import Font, PatternFill, Alignment

# --- Schema Art. 2424 (Stato Patrimoniale) ---
ART_2424_ATTIVO = [
    ("A", "Crediti verso soci per versamenti ancora dovuti", True),
    ("B", "Immobilizzazioni", True),
    ("B.I", "Immobilizzazioni immateriali", True),
    ("B.I.1", "costi di impianto e di ampliamento", False),
    ("B.I.2", "costi di sviluppo", False),
    ("B.I.3", "diritti di brevetto industriale e diritti di utilizzazione delle opere dell'ingegno", False),
    ("B.I.4", "concessioni, licenze, marchi e diritti simili", False),
    ("B.I.5", "avviamento", False),
    ("B.I.6", "immobilizzazioni in corso e acconti", False),
    ("B.I.7", "altre", False),
    ("B.I_Tot", "Totale", True),
    ("B.II", "Immobilizzazioni materiali", True),
    ("B.II.1", "terreni e fabbricati", False),
    ("B.II.2", "impianti e macchinario", False),
    ("B.II.3", "attrezzature industriali e commerciali", False),
    ("B.II.4", "altri beni", False),
    ("B.II.5", "immobilizzazioni in corso e acconti", False),
    ("B.II_Tot", "Totale", True),
    ("B.III", "Immobilizzazioni finanziarie", True),
    ("B.III.1", "partecipazioni", False),
    ("B.III.2", "crediti", False),
    ("B.III.3", "altri titoli", False),
    ("B.III.4", "strumenti finanziari derivati attivi", False),
    ("B.III_Tot", "Totale", True),
    ("B_Tot", "Totale immobilizzazioni", True),
    ("C", "Attivo circolante", True),
    ("C.I", "Rimanenze", True),
    ("C.I.1", "materie prime, sussidiarie e di consumo", False),
    ("C.I.2", "prodotti in corso di lavorazione e semilavorati", False),
    ("C.I.3", "lavori in corso su ordinazione", False),
    ("C.I.4", "prodotti finiti e merci", False),
    ("C.I.5", "acconti", False),
    ("C.I_Tot", "Totale", True),
    ("C.II", "Crediti", True),
    ("C.II.1", "verso clienti", False),
    ("C.II.2", "verso imprese controllate", False),
    ("C.II.3", "verso imprese collegate", False),
    ("C.II.4", "verso controllanti", False),
    ("C.II.5", "verso imprese sottoposte al controllo delle controllanti", False),
    ("C.II.5-bis", "crediti tributari", False),
    ("C.II.5-ter", "imposte anticipate", False),
    ("C.II.5-quater", "verso altri", False),
    ("C.II_Tot", "Totale", True),
    ("C.III", "Attività finanziarie che non costituiscono immobilizzazioni", True),
    ("C.III.1", "partecipazioni in imprese controllate", False),
    ("C.III.2", "partecipazioni in imprese collegate", False),
    ("C.III.3", "partecipazioni in imprese controllanti", False),
    ("C.III.3-bis", "partecipazioni in imprese sottoposte al controllo", False),
    ("C.III.4", "altre partecipazioni", False),
    ("C.III.5", "strumenti finanziari derivati attivi", False),
    ("C.III.6", "altri titoli", False),
    ("C.III_Tot", "Totale", True),
    ("C.IV", "Disponibilità liquide", True),
    ("C.IV.1", "depositi bancari e postali", False),
    ("C.IV.2", "assegni", False),
    ("C.IV.3", "danaro e valori in cassa", False),
    ("C.IV_Tot", "Totale", True),
    ("C_Tot", "Totale attivo circolante", True),
    ("D", "Ratei e risconti", True),
    ("TOT_ATTIVO", "TOTALE ATTIVO", True)
]

ART_2424_PASSIVO = [
    ("A", "Patrimonio netto", True),
    ("A.I", "Capitale", False),
    ("A.II", "Riserva da soprapprezzo delle azioni", False),
    ("A.III", "Riserve di rivalutazione", False),
    ("A.IV", "Riserva legale", False),
    ("A.V", "Riserve statutarie", False),
    ("A.VI", "Altre riserve", False),
    ("A.VII", "Riserva per operazioni di copertura dei flussi finanziari attesi", False),
    ("A.VIII", "Utili (perdite) portato a nuovo", False),
    ("A.IX", "Utile (perdita) dell'esercizio", False),
    ("A.X", "Riserva negativa per azioni proprie in portafoglio", False),
    ("A_Tot", "Totale Patrimonio netto", True),
    ("B", "Fondi per rischi e oneri", True),
    ("B.1", "per trattamento di quiescenza e obblighi simili", False),
    ("B.2", "per imposte, anche differite", False),
    ("B.3", "strumenti finanziari derivati passivi", False),
    ("B.4", "altri", False),
    ("B_Tot", "Totale", True),
    ("C", "Trattamento di fine rapporto di lavoro subordinato", True),
    ("D", "Debiti", True),
    ("D.1", "obbligazioni", False),
    ("D.2", "obbligazioni convertibili", False),
    ("D.3", "debiti verso soci per finanziamenti", False),
    ("D.4", "debiti verso banche", False),
    ("D.5", "debiti verso altri finanziatori", False),
    ("D.6", "acconti", False),
    ("D.7", "debiti verso fornitori", False),
    ("D.8", "debiti rappresentati da titoli di credito", False),
    ("D.9", "debiti verso imprese controllate", False),
    ("D.10", "debiti verso imprese collegate", False),
    ("D.11", "debiti verso controllanti", False),
    ("D.11-bis", "debiti verso imprese sottoposte al controllo delle controllanti", False),
    ("D.12", "debiti tributari", False),
    ("D.13", "debiti verso istituti di previdenza e di sicurezza sociale", False),
    ("D.14", "altri debiti", False),
    ("D_Tot", "Totale", True),
    ("E", "Ratei e risconti", True),
    ("TOT_PASSIVO", "TOTALE PASSIVO", True)
]

# --- Schema Art. 2425 (Conto Economico) ---
ART_2425 = [
    ("A", "Valore della produzione:", True),
    ("A.1", "ricavi delle vendite e delle prestazioni", False),
    ("A.2", "variazioni delle rimanenze di prodotti in corso di lavorazione, semilavorati e finiti", False),
    ("A.3", "variazioni dei lavori in corso su ordinazione", False),
    ("A.4", "incrementi di immobilizzazioni per lavori interni", False),
    ("A.5", "altri ricavi e proventi", False),
    ("A_Tot", "Totale Valore della produzione", True),
    ("B", "Costi della produzione:", True),
    ("B.6", "per materie prime, sussidiarie, di consumo e di merci", False),
    ("B.7", "per servizi", False),
    ("B.8", "per godimento di beni di terzi", False),
    ("B.9", "per il personale", True),
    ("B.9.a", "salari e stipendi", False),
    ("B.9.b", "oneri sociali", False),
    ("B.9.c", "trattamento di fine rapporto", False),
    ("B.9.d", "trattamento di quiescenza e simili", False),
    ("B.9.e", "altri costi", False),
    ("B.10", "ammortamenti e svalutazioni", True),
    ("B.10.a", "ammortamento delle immobilizzazioni immateriali", False),
    ("B.10.b", "ammortamento delle immobilizzazioni materiali", False),
    ("B.10.c", "altre svalutazioni delle immobilizzazioni", False),
    ("B.10.d", "svalutazioni dei crediti", False),
    ("B.11", "variazioni delle rimanenze", False),
    ("B.12", "accantonamenti per rischi", False),
    ("B.13", "altri accantonamenti", False),
    ("B.14", "oneri diversi di gestione", False),
    ("B_Tot", "Totale Costi della produzione", True),
    ("A_B_Diff", "Differenza tra valore e costi della produzione (A - B)", True),
    ("C", "Proventi e oneri finanziari:", True),
    ("C.15", "proventi da partecipazioni", False),
    ("C.16", "altri proventi finanziari", True),
    ("C.16.a", "da crediti iscritti nelle immobilizzazioni", False),
    ("C.16.b", "da titoli iscritti nelle immobilizzazioni", False),
    ("C.16.c", "da titoli iscritti nell'attivo circolante", False),
    ("C.16.d", "proventi diversi dai precedenti", False),
    ("C.17", "interessi e altri oneri finanziari", False),
    ("C.17-bis", "utili e perdite su cambi", False),
    ("C_Tot", "Totale (15 + 16 - 17 +/- 17bis)", True),
    ("D", "Rettifiche di valore di attività e passività finanziarie:", True),
    ("D.18", "rivalutazioni", True),
    ("D.18.a", "di partecipazioni", False),
    ("D.18.b", "di immobilizzazioni finanziarie", False),
    ("D.18.c", "di titoli iscritti all'attivo circolante", False),
    ("D.18.d", "di strumenti finanziari derivati", False),
    ("D.19", "svalutazioni", True),
    ("D.19.a", "di partecipazioni", False),
    ("D.19.b", "di immobilizzazioni finanziarie", False),
    ("D.19.c", "di titoli iscritti nell'attivo circolante", False),
    ("D.19.d", "di strumenti finanziari derivati", False),
    ("D_Tot", "Totale delle rettifiche (18-19)", True),
    ("E", "Risultato prima delle imposte (A - B +/- C +/- D)", True),
    ("20", "imposte sul reddito dell'esercizio", False),
    ("21", "utile (perdita) dell'esercizio", True)
]

def load_mapping(sys_dir):
    mapping_path = os.path.join(sys_dir, "Bilanz_Mapping.xlsx")
    mapping = {}
    if os.path.exists(mapping_path):
        try:
            df_map = pd.read_excel(mapping_path)
            for _, row in df_map.iterrows():
                conto = str(row.get('Conto', '')).strip()
                pos = str(row.get('Art_2424_Position', row.get('Position', ''))).strip()
                if conto and pos and pos != 'nan':
                    mapping[conto] = pos
        except Exception as e:
            print(f"Warnung: Konnte Bilanz_Mapping.xlsx nicht laden: {e}")
    return mapping

def generate_bilanz_worksheet(writer, alle_positionen, sys_dir):
    mapping = load_mapping(sys_dir)
    
    bal_2424 = {k: 0.0 for k, _, _ in ART_2424_ATTIVO + ART_2424_PASSIVO}
    bal_2425 = {k: 0.0 for k, _, _ in ART_2425}
    
    attivo_keys = [k for k, _, _ in ART_2424_ATTIVO]
    passivo_keys = [k for k, _, _ in ART_2424_PASSIVO]
    
    # 2425: Elements that increase the Utile
    utile_increasing = ["A.1", "A.2", "A.3", "A.4", "A.5", "C.15", "C.16.a", "C.16.b", "C.16.c", "C.16.d", "C.17-bis", "D.18.a", "D.18.b", "D.18.c", "D.18.d"]
    
    for pos in alle_positionen:
        aktiv_passiv = pos.get('Aktiv/Passiv', '')
        if not aktiv_passiv:
            continue
            
        conto_val = pos.get('Conto', '')
        conto_str = str(conto_val).strip() if conto_val else ''
        
        gesamt_key = next((k for k in pos.keys() if 'Gesamtpreis' in k), None)
        netto = float(pos.get(gesamt_key, 0.0) or 0.0)
        
        mwst_rate = float(pos.get('MwSt (%)', 0.0) or 0.0)
        steuer = netto * mwst_rate
        brutto = netto + steuer
        
        mapped_pos = mapping.get(conto_str, "")
        
        if aktiv_passiv == 'Attiva':
            # Crediti verso clienti
            bal_2424["C.II.1"] = bal_2424.get("C.II.1", 0.0) + brutto
            # IVA a debito -> Debiti tributari
            bal_2424["D.12"] = bal_2424.get("D.12", 0.0) + steuer
            
            # Distribuzione del Netto
            if mapped_pos in bal_2425:
                if mapped_pos in utile_increasing:
                    bal_2425[mapped_pos] += netto
                else:
                    bal_2425[mapped_pos] -= netto
            elif mapped_pos in attivo_keys:
                bal_2424[mapped_pos] -= netto
            elif mapped_pos in passivo_keys:
                bal_2424[mapped_pos] += netto
            else:
                bal_2425["A.5"] += netto # Unmapped ricavi
                
        elif aktiv_passiv == 'Passiva':
            # Debiti verso fornitori
            bal_2424["D.7"] = bal_2424.get("D.7", 0.0) + brutto
            # IVA a credito -> Crediti tributari
            bal_2424["C.II.5-bis"] = bal_2424.get("C.II.5-bis", 0.0) + steuer
            
            # Distribuzione del Netto
            if mapped_pos in bal_2425:
                if mapped_pos in utile_increasing:
                    bal_2425[mapped_pos] -= netto
                else:
                    bal_2425[mapped_pos] += netto
            elif mapped_pos in attivo_keys:
                bal_2424[mapped_pos] += netto
            elif mapped_pos in passivo_keys:
                bal_2424[mapped_pos] -= netto
            else:
                bal_2425["B.14"] += netto # Unmapped costi

    # --- Calculate Subtotals 2425 ---
    bal_2425["A_Tot"] = sum(bal_2425.get(f"A.{i}", 0.0) for i in range(1, 6))
    
    bal_2425["B.9"] = sum(bal_2425.get(f"B.9.{x}", 0.0) for x in ["a", "b", "c", "d", "e"])
    bal_2425["B.10"] = sum(bal_2425.get(f"B.10.{x}", 0.0) for x in ["a", "b", "c", "d"])
    bal_2425["B_Tot"] = sum(bal_2425.get(f"B.{i}", 0.0) for i in [6, 7, 8, 9, 10, 11, 12, 13, 14])
    
    bal_2425["A_B_Diff"] = bal_2425["A_Tot"] - bal_2425["B_Tot"]
    
    bal_2425["C.16"] = sum(bal_2425.get(f"C.16.{x}", 0.0) for x in ["a", "b", "c", "d"])
    bal_2425["C_Tot"] = bal_2425.get("C.15", 0.0) + bal_2425["C.16"] - bal_2425.get("C.17", 0.0) + bal_2425.get("C.17-bis", 0.0)
    
    bal_2425["D.18"] = sum(bal_2425.get(f"D.18.{x}", 0.0) for x in ["a", "b", "c", "d"])
    bal_2425["D.19"] = sum(bal_2425.get(f"D.19.{x}", 0.0) for x in ["a", "b", "c", "d"])
    bal_2425["D_Tot"] = bal_2425["D.18"] - bal_2425["D.19"]
    
    bal_2425["E"] = bal_2425["A_B_Diff"] + bal_2425["C_Tot"] + bal_2425["D_Tot"]
    bal_2425["21"] = bal_2425["E"] - bal_2425.get("20", 0.0)
    
    # --- Inject Utile into 2424 ---
    bal_2424["A.IX"] = bal_2425["21"]

    # --- Calculate Subtotals 2424 ---
    bal_2424["B.I_Tot"] = sum(bal_2424.get(f"B.I.{i}", 0.0) for i in range(1, 8))
    bal_2424["B.II_Tot"] = sum(bal_2424.get(f"B.II.{i}", 0.0) for i in range(1, 6))
    bal_2424["B.III_Tot"] = sum(bal_2424.get(f"B.III.{i}", 0.0) for i in range(1, 5))
    bal_2424["B_Tot"] = bal_2424["B.I_Tot"] + bal_2424["B.II_Tot"] + bal_2424["B.III_Tot"]
    
    bal_2424["C.I_Tot"] = sum(bal_2424.get(f"C.I.{i}", 0.0) for i in range(1, 6))
    bal_2424["C.II_Tot"] = sum(bal_2424.get(f"C.II.{i}", 0.0) for i in [1, 2, 3, 4, 5, "5-bis", "5-ter", "5-quater"])
    bal_2424["C.III_Tot"] = sum(bal_2424.get(f"C.III.{i}", 0.0) for i in [1, 2, 3, "3-bis", 4, 5, 6])
    bal_2424["C.IV_Tot"] = sum(bal_2424.get(f"C.IV.{i}", 0.0) for i in range(1, 4))
    bal_2424["C_Tot"] = bal_2424["C.I_Tot"] + bal_2424["C.II_Tot"] + bal_2424["C.III_Tot"] + bal_2424["C.IV_Tot"]
    
    bal_2424["TOT_ATTIVO"] = bal_2424.get("A", 0.0) + bal_2424["B_Tot"] + bal_2424["C_Tot"] + bal_2424.get("D", 0.0)
    
    bal_2424["A_Tot"] = sum(bal_2424.get(f"A.{r}", 0.0) for r in ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"])
    bal_2424["B_Tot"] = sum(bal_2424.get(f"B.{i}", 0.0) for i in range(1, 5))
    bal_2424["D_Tot"] = sum(bal_2424.get(f"D.{i}", 0.0) for i in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, "11-bis", 12, 13, 14])
    
    bal_2424["TOT_PASSIVO"] = bal_2424["A_Tot"] + bal_2424["B_Tot"] + bal_2424.get("C", 0.0) + bal_2424["D_Tot"] + bal_2424.get("E", 0.0)
    
    wb = writer.book
    
    # If the default "Bilanz" sheet exists, remove it
    if "Bilanz" in wb.sheetnames:
        del wb["Bilanz"]
    
    # ---------------------------------------------
    # Write 2424 Sheet
    # ---------------------------------------------
    ws_2424 = wb.create_sheet("2424")
    
    ws_2424.merge_cells('A1:F1')
    warning_cell = ws_2424.cell(row=1, column=1, value="⚠️ Entwurf: Saldi der liquiden Mittel e Zahlungsabgleiche fehlen")
    warning_cell.font = Font(color="FF0000", bold=True, size=12)
    warning_cell.alignment = Alignment(horizontal="center")
    warning_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    ws_2424.cell(row=3, column=1, value="ATTIVO").font = Font(bold=True, size=14)
    ws_2424.cell(row=3, column=4, value="PASSIVO").font = Font(bold=True, size=14)
    
    current_row_attivo = 4
    for code, desc, is_bold in ART_2424_ATTIVO:
        val = bal_2424.get(code, 0.0)
        c_code = ws_2424.cell(row=current_row_attivo, column=1, value=code)
        c_desc = ws_2424.cell(row=current_row_attivo, column=2, value=desc)
        c_val = ws_2424.cell(row=current_row_attivo, column=3, value=val)
        
        c_val.number_format = '#,##0.00 €'
        if is_bold:
            c_code.font = Font(bold=True)
            c_desc.font = Font(bold=True)
            c_val.font = Font(bold=True)
        current_row_attivo += 1
        
    current_row_passivo = 4
    for code, desc, is_bold in ART_2424_PASSIVO:
        val = bal_2424.get(code, 0.0)
        c_code = ws_2424.cell(row=current_row_passivo, column=4, value=code)
        c_desc = ws_2424.cell(row=current_row_passivo, column=5, value=desc)
        c_val = ws_2424.cell(row=current_row_passivo, column=6, value=val)
        
        c_val.number_format = '#,##0.00 €'
        if is_bold:
            c_code.font = Font(bold=True)
            c_desc.font = Font(bold=True)
            c_val.font = Font(bold=True)
        current_row_passivo += 1
        
    ws_2424.column_dimensions['A'].width = 15
    ws_2424.column_dimensions['B'].width = 60
    ws_2424.column_dimensions['C'].width = 18
    ws_2424.column_dimensions['D'].width = 15
    ws_2424.column_dimensions['E'].width = 60
    ws_2424.column_dimensions['F'].width = 18

    # ---------------------------------------------
    # Write 2425 Sheet
    # ---------------------------------------------
    ws_2425 = wb.create_sheet("2425")
    
    ws_2425.cell(row=2, column=1, value="CONTO ECONOMICO (Art. 2425)").font = Font(bold=True, size=14)
    
    current_row_2425 = 4
    for code, desc, is_bold in ART_2425:
        val = bal_2425.get(code, 0.0)
        c_code = ws_2425.cell(row=current_row_2425, column=1, value=code)
        c_desc = ws_2425.cell(row=current_row_2425, column=2, value=desc)
        c_val = ws_2425.cell(row=current_row_2425, column=3, value=val)
        
        c_val.number_format = '#,##0.00 €'
        if is_bold:
            c_code.font = Font(bold=True)
            c_desc.font = Font(bold=True)
            c_val.font = Font(bold=True)
            
        current_row_2425 += 1
        
    ws_2425.column_dimensions['A'].width = 15
    ws_2425.column_dimensions['B'].width = 80
    ws_2425.column_dimensions['C'].width = 18
