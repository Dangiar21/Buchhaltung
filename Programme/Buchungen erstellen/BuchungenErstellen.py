import sys
import os
import traceback
import re
from difflib import SequenceMatcher

# Utils aus dem übergeordneten Ordner laden
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass
from utils import load_or_create_targa_list, append_new_targas_to_excel, ask_shorten_desc, get_text, safe_float, read_xml_or_p7m, is_similar_desc, is_generic_auxiliary

# 1. Wir versuchen die Module zu laden. Wenn das fehlschlägt, fangen wir den Fehler ab.
import Buchung_Regeln
try:
    import defusedxml.ElementTree as ET
    import pandas as pd
except Exception as e:
    print(f"Fehler beim Laden der Module: {e}")
    print("Hast du 'pip install pandas openpyxl' im Terminal ausgeführt?")
    input("\nDrücke Enter zum Beenden...")
    sys.exit(1)


from sdi_parser import parse_sdi_xml

def clean_description_for_dedup(desc):
    if not desc: return ""
    desc_upper = desc.upper()
    
    # 1. Sicherer Fallback aus der Originalbeschreibung (bereinigt von wirren Steuerzeichen)
    original_for_fallback = re.sub(r'[\r\n\t]', ' ', desc_upper)
    original_for_fallback = re.sub(r'[^\w\s.,-]', ' ', original_for_fallback)
    original_for_fallback = re.sub(r'\s+', ' ', original_for_fallback).strip()
    
    # 2. Datumsmuster entfernen (DD.MM.YYYY, YYYY-MM-DD, DD/MM/YY, DD.MM.YY)
    cleaned = re.sub(r'\b(\d{1,2}[\./-]\d{1,2}[\./-]\d{2,4}|\d{4}[\./-]\d{1,2}[\./-]\d{1,2})\b', ' ', desc_upper)
    
    # 3. Monatsnamen (DE & IT) mit optionalem Jahr entfernen
    months_de = r'JANUAR|FEBRUAR|MÄRZ|APRIL|MAI|JUNI|JULI|AUGUST|SEPTEMBER|OKTOBER|NOVEMBER|DEZEMBER|JAN|FEB|MÄR|APR|JUN|JUL|AUG|SEP|OKT|NOV|DEZ'
    months_it = r'GENNAIO|FEBBRAIO|MARZO|APRILE|MAGGIO|GIUGNO|LUGLIO|AGOSTO|SETTEMBRE|OTTOBRE|NOVEMBRE|DICEMBRE|GEN|MAG|GIU|LUG|AGO|SET|OTT|DIC'
    months = rf'\b({months_de}|{months_it})\b'
    cleaned = re.sub(months + r'(\s*\d{2,4})?', ' ', cleaned)
    
    # 4. Mengenangaben mit expliziten Maßeinheiten entfernen (z.B. 500 KG, 290ML, 10 STK, 0,6 MM)
    cleaned = re.sub(r'\b\d+([.,]\d+)*\s*(KG|G|L|LT|ML|CM|MM|M|STK|STÜCK|PZ|%)\b', ' ', cleaned)
    
    # 5. Spezifische Tier-Ohrmarken / amtliche IDs gezielt entfernen (z.B. IT021000123456, DE0912345678)
    cleaned = re.sub(r'\b(IT|DE|AT|FR|NL)\s*\d{7,14}\b', ' ', cleaned)
    
    # 6. Explizite Seriennummer-Präfixe entfernen (z.B. S/N: 12345, MATR. 9482)
    cleaned = re.sub(r'\b(S/N|SN|MATR|MATRICOLA)[\s.:]*[A-Z0-9-]+\b', ' ', cleaned)
    
    # 7. Reine Jahreszahlen am Textende entfernen (z.B. "BEITRAG 2023" -> "BEITRAG")
    cleaned = re.sub(r'\b(19\d{2}|20\d{2})\s*$', ' ', cleaned)
    
    # 8. Satzzeichen & Trennzeichen (Klammern, Schrägstriche, Bindestriche etc. in Leerzeichen)
    cleaned = re.sub(r'[()/\-:;\[\]{}+*?!~"\'`_]', ' ', cleaned)
    # Entferne isolierte Punkte und Kommas, die nicht von Ziffern umgeben sind
    cleaned = re.sub(r'(?<!\d)[.,]|[.,](?!\d)', ' ', cleaned)
    
    # 9. Mehrfache Leerzeichen zusammenfassen
    cleaned = re.sub(r'\s+', ' ', cleaned).strip(' ,.-')
    
    # 10. SICHERHEITS-GUARD / FALLBACK:
    # Wenn weniger als 2 Buchstaben übrig sind, sofort Fallback auf das gereinigte Original
    letters = re.findall(r'[A-ZÄÖÜa-zäöü]', cleaned)
    if len(letters) < 2:
        orig_letters = re.findall(r'[A-ZÄÖÜa-zäöü]', original_for_fallback)
        if len(orig_letters) >= 2:
            return original_for_fallback
        elif original_for_fallback:
            return original_for_fallback
        return desc.strip()
        
    return cleaned

def parse_xml_to_list(xml_path, targa_dict, neue_targas_set, fehler_log, rules_dict, shorten_description=False, client_vat_id="", db_konten_cache=None):
    if db_konten_cache is None: db_konten_cache = {}
    print(f"Lese: {xml_path}")
    
    parsed_items = parse_sdi_xml(xml_path, targa_dict, neue_targas_set, fehler_log, shorten_description, client_vat_id)
    rechnungspositionen = []
    
    # 1. Hauptleistung der Rechnung ermitteln (Sachposition mit höchstem Gesamtpreis, die kein reiner Nebenkostenbegriff ist)
    invoice_main_desc = ""
    invoice_main_item = None
    max_amount = -1.0
    for it in parsed_items:
        it_desc = it.get('Beschreibung', '').strip()
        it_clean = clean_description_for_dedup(it_desc)
        it_amount = abs(safe_float(str(it.get('Gesamtpreis_Roh', 0.0)), 0.0))
        if not is_generic_auxiliary(it_clean) and it_amount > max_amount and it_clean:
            max_amount = it_amount
            invoice_main_desc = it_clean
            invoice_main_item = it
            
    # Falls alle Positionen generisch sind (z.B. reine Gebührenrechnung), nimm die größte Position als Fallback
    if not invoice_main_desc and parsed_items:
        for it in parsed_items:
            it_desc = it.get('Beschreibung', '').strip()
            it_clean = clean_description_for_dedup(it_desc)
            it_amount = abs(safe_float(str(it.get('Gesamtpreis_Roh', 0.0)), 0.0))
            if it_amount > max_amount and it_clean:
                max_amount = it_amount
                invoice_main_desc = it_clean
                invoice_main_item = it

    # 2. Konto der Hauptleistung vorab ermitteln (falls bereits in Regeln oder Cache vorhanden)
    main_conto = "???"
    main_is_pending = True
    if invoice_main_item:
        main_clean = clean_description_for_dedup(invoice_main_item.get('Beschreibung', ''))
        # a) Regeln
        m_conto, m_pending = Buchung_Regeln.assign_account(
            invoice_main_item['Desc_Norm'], invoice_main_item['Beschreibung'],
            invoice_main_item['Lieferant'], invoice_main_item['Liefer ID'], invoice_main_item['Kunden ID'],
            rules_dict
        )
        if m_conto != "???":
            main_conto = m_conto
            main_is_pending = m_pending
        else:
            # b) Cache
            main_cache_key = f"{invoice_main_item['Lieferant']} | {main_clean}".strip().upper()
            main_cache_raw = f"{invoice_main_item['Lieferant']} | {invoice_main_item['Beschreibung']}".strip().upper()
            if main_cache_key in db_konten_cache:
                main_conto = str(db_konten_cache[main_cache_key]['value'])
                main_is_pending = not db_konten_cache[main_cache_key]['confirmed']
            elif main_cache_raw in db_konten_cache:
                main_conto = str(db_konten_cache[main_cache_raw]['value'])
                main_is_pending = not db_konten_cache[main_cache_raw]['confirmed']
            else:
                m_match = find_fuzzy_cache_match(invoice_main_item['Lieferant'], main_clean, db_konten_cache)
                if m_match:
                    main_conto = str(m_match['value'])
                    main_is_pending = not m_match['confirmed']
    
    for item in parsed_items:
        clean_desc = clean_description_for_dedup(item['Beschreibung'])
        is_aux = is_generic_auxiliary(clean_desc)
        rechnung_kontext = invoice_main_desc if (is_aux and invoice_main_desc and invoice_main_desc != clean_desc) else ""

        # Konto ermitteln
        # 1. Priorität: Benutzer- & Globale Kontenregeln (Kunde vor Global)
        conto, is_pending = Buchung_Regeln.assign_account(
            item['Desc_Norm'], item['Beschreibung'], item['Lieferant'], item['Liefer ID'], item['Kunden ID'], rules_dict
        )
        
        # 2. Priorität: Falls keine Regel greift, im Datenbank-Cache suchen oder von Hauptleistung erben
        if conto == "???":
            if is_aux:
                # Bei Nebenpositionen (Netzausgaben, Steuern/Gebühren, Spesen):
                # Prüfe zuerst, ob genau dieser Kontext schon im Cache liegt
                cache_key_ctx = f"{item['Lieferant']} | {clean_desc} [KONTEXT: {rechnung_kontext}]".strip().upper()
                if rechnung_kontext and cache_key_ctx in db_konten_cache:
                    conto = str(db_konten_cache[cache_key_ctx]['value'])
                    is_pending = not db_konten_cache[cache_key_ctx]['confirmed']
                elif main_conto != "???":
                    # Vererbung: Nebenpositionen teilen das Konto der Hauptleistung (z.B. Gas -> 810004)!
                    conto = main_conto
                    is_pending = main_is_pending
                    if rechnung_kontext:
                        db_konten_cache[cache_key_ctx] = {'value': main_conto, 'confirmed': False}
                # Kein Fallback auf cache_key_raw bei Nebenpositionen (verhindert Vermischung von Gas & Strom!)
            else:
                # Haupt- / Sachposition
                cache_key = f"{item['Lieferant']} | {clean_desc}".strip().upper()
                cache_key_raw = f"{item['Lieferant']} | {item['Beschreibung']}".strip().upper()
                
                if cache_key in db_konten_cache:
                    conto = str(db_konten_cache[cache_key]['value'])
                    is_pending = not db_konten_cache[cache_key]['confirmed']
                elif cache_key_raw in db_konten_cache:
                    conto = str(db_konten_cache[cache_key_raw]['value'])
                    is_pending = not db_konten_cache[cache_key_raw]['confirmed']
                else:
                    # Fuzzy Cache Match (gleicher Lieferant, ähnliche Beschreibung unter Beachtung von Signalwörtern)
                    matched_cache = find_fuzzy_cache_match(item['Lieferant'], clean_desc, db_konten_cache)
                    if matched_cache:
                        conto = str(matched_cache['value'])
                        is_pending = not matched_cache['confirmed']
                        db_konten_cache[cache_key] = matched_cache
                        if clean_desc == invoice_main_desc and main_conto == "???":
                            main_conto = conto
                            main_is_pending = is_pending
            
        waehrung = item.get('Waehrung', 'EUR')
        
        rechnungspositionen.append({
            'Aktiv/Passiv': item['Aktiv/Passiv'],
            'Typ': item['Typ'],
            'Rechnungsnummer': item['Rechnungsnummer'],
            'Datum': item['Datum'],
            'Lieferant': item['Lieferant'],
            'Liefer ID': item['Liefer ID'],
            'Kunde': item['Kunde'],
            'Kunden ID': item['Kunden ID'],
            'Beschreibung': item['Beschreibung'],
            'Unterkonto': conto,
            'Hauptkonto': '',
            'is_pending': is_pending,
            '_rechnung_kontext': rechnung_kontext,
            'CdC': item['CdC'],
            'Kennzeichen': item['Kennzeichen'],
            'Fahrzeugtyp': item['Fahrzeugtyp'],
            'Menge': item['Menge'],
            f'Einzelpreis ({waehrung})': item['Einzelpreis_Roh'],
            f'Gesamtpreis ({waehrung})': item['Gesamtpreis_Roh'],
            'MwSt (%)': item['MwSt']
        })
        
    # Falls das Konto der Hauptleistung erst im Verlauf der Schleife bekannt wurde,
    # nachträglich alle noch offenen Nebenpositionen mit main_conto aktualisieren
    if main_conto != "???":
        for pos in rechnungspositionen:
            if pos['Unterkonto'] == "???" and pos.get('_rechnung_kontext'):
                pos['Unterkonto'] = main_conto
                pos['is_pending'] = main_is_pending
                c_clean = clean_description_for_dedup(pos['Beschreibung'])
                ck = f"{pos['Lieferant']} | {c_clean} [KONTEXT: {pos['_rechnung_kontext']}]".strip().upper()
                db_konten_cache[ck] = {'value': main_conto, 'confirmed': False}
        
    return rechnungspositionen

def find_fuzzy_cache_match(supplier_name, desc_to_match, db_konten_cache, threshold=0.80):
    """
    Sucht im Konten-Cache nach einem Eintrag desselben Lieferanten mit hinreichend ähnlicher Beschreibung.
    Gibt das gefundene Dict {'value': konto, 'confirmed': bool} oder None zurück.
    """
    if not supplier_name or not desc_to_match or not db_konten_cache:
        return None
    # Nebenpositionen dürfen niemals generisch per Fuzzy gematcht werden, da sie vom Kontext abhängen
    if is_generic_auxiliary(desc_to_match):
        return None

    supplier_upper = str(supplier_name).strip().upper()
    desc_upper = str(desc_to_match).strip().upper()
    best_match = None
    highest_ratio = 0.0

    for key, data in db_konten_cache.items():
        if ' | ' in key:
            k_supp, k_desc = key.split(' | ', 1)
            k_supp = k_supp.strip()
            k_desc = k_desc.strip()
            k_desc_clean = re.sub(r'\s*\[KONTEXT:.*?\]', '', k_desc, flags=re.IGNORECASE).strip()
            if k_supp == supplier_upper or (len(k_supp) >= 5 and (k_supp in supplier_upper or supplier_upper in k_supp)):
                if not k_desc_clean:
                    continue
                ratio = SequenceMatcher(None, desc_upper, k_desc_clean).ratio()
                if is_similar_desc(desc_upper, k_desc_clean, threshold=threshold):
                    effective_score = max(ratio, threshold)
                    if effective_score > highest_ratio:
                        highest_ratio = effective_score
                        best_match = data
    return best_match

def run_conversion(paths=None, output_dir=None, nutzerdaten_dir=None):
    if paths is None:
        paths = sys.argv[1:]
        
    alle_positionen = []
    ausgabe_ordner = output_dir

    try:
        if len(paths) > 0:
            for pfad in paths:
                # Setze den Ausgabeordner auf das Verzeichnis des ersten Elements, falls keiner gegeben
                if not ausgabe_ordner:
                    if os.path.isfile(pfad):
                        ausgabe_ordner = os.path.dirname(pfad)
                    else:
                        ausgabe_ordner = pfad

            # Lade oder erstelle die Targa Liste VOR dem Parsen der XML Dateien
            targa_dict, targa_file = load_or_create_targa_list(nutzerdaten_dir)
            
            # --- Regel-System initialisieren ---
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            
            global_rules_path = os.path.join(base_dir, "Systemdaten", "Globale_KontenRegeln.xlsx")
            Buchung_Regeln.ensure_rule_file(global_rules_path)
            
            client_rules_path = None
            if nutzerdaten_dir:
                client_rules_path = os.path.join(nutzerdaten_dir, "Kunden_KontenRegeln.xlsx")
                Buchung_Regeln.ensure_rule_file(client_rules_path)
            else:
                client_rules_path = os.path.join(base_dir, "Kunden", "Unbekannt", "Nutzerdaten", "Kunden_KontenRegeln.xlsx")
                Buchung_Regeln.ensure_rule_file(client_rules_path)
                
            rules_dict = Buchung_Regeln.load_rules(global_rules_path, client_rules_path)
            
            neue_targas_set = set()
            fehler_log = []
            
            client_vat_id = ""
            client_name = os.path.basename(os.path.dirname(nutzerdaten_dir)) if nutzerdaten_dir else "Unbekannt"
            if client_name != "Unbekannt":
                try:
                    from src.db.database import init_db, Kunde
                    db_path = os.path.join(base_dir, "Kunden", "kunden.db")
                    session = init_db(db_path)
                    kunde = session.query(Kunde).filter_by(name=client_name).first()
                    if kunde:
                        client_vat_id = (kunde.partita_iva or "").strip()
                        if not client_vat_id:
                            client_vat_id = (kunde.codice_fiscale or "").strip()
                except Exception as e:
                    print(f"Fehler beim Lesen der Datenbank: {e}")
            
            # Lade den DB Cache für den Kunden, um UI-Bestätigungen zu berücksichtigen
            kunden_id_ordner = os.path.basename(os.path.dirname(nutzerdaten_dir)) if nutzerdaten_dir else "Unbekannt"
            try:
                from DatabaseManager import get_db
                db = get_db()
                db_konten_cache = db.get_konten_cache_full(kunden_id_ordner)
            except Exception:
                db_konten_cache = {}
            
            shorten_description = ask_shorten_desc()

            xml_files_to_process = []
            for pfad in paths:
                if os.path.isfile(pfad):
                    lower_pfad = pfad.lower()
                    if lower_pfad.endswith('.xml') or lower_pfad.endswith('.p7m'):
                        xml_files_to_process.append(pfad)
                elif os.path.isdir(pfad):
                    print(f"\nDurchsuche Ordner (inkl. Unterordner): {pfad}")
                    for root_dir, _, files in os.walk(pfad):
                        for filename in files:
                            lower_file = filename.lower()
                            if lower_file.endswith('.xml') or lower_file.endswith('.p7m'):
                                xml_files_to_process.append(os.path.join(root_dir, filename))
                else:
                    print(f"Überspringe: {pfad} (Keine XML/P7M oder Ordner)")
                    
            total_files = len(xml_files_to_process)
            for i, xml_file in enumerate(xml_files_to_process):
                alle_positionen.extend(parse_xml_to_list(xml_file, targa_dict, neue_targas_set, fehler_log, rules_dict, shorten_description, client_vat_id, db_konten_cache))
                percent = int(((i + 1) / total_files) * 20) if total_files > 0 else 20
                print(f"[PROGRESS:{percent}]")
                    
            
            if alle_positionen:
                # --- Heuristik für Aktiv/Passiv (verschoben vor KI) ---
                all_vats = []
                for pos in alle_positionen:
                    if pos.get('Liefer ID'): all_vats.append(pos['Liefer ID'])
                    if pos.get('Kunden ID'): all_vats.append(pos['Kunden ID'])
                if all_vats:
                    from collections import Counter
                    guessed_vat = Counter(all_vats).most_common(1)[0][0]
                    for pos in alle_positionen:
                        if not pos.get('Aktiv/Passiv'):
                            if pos.get('Liefer ID') == guessed_vat:
                                pos['Aktiv/Passiv'] = 'Attiva'
                            elif pos.get('Kunden ID') == guessed_vat:
                                pos['Aktiv/Passiv'] = 'Passiva'
                            else:
                                pos['Aktiv/Passiv'] = 'Passiva'
                                
                # --- KI Fallback ---
                import Buchung_KI
                api_key = Buchung_KI.get_api_key(base_dir)
                
                unique_unknowns_er = {}
                unique_unknowns_ar = {}
                
                for i, pos in enumerate(alle_positionen):
                    if pos.pop('is_pending', False):
                        pos['_is_ai'] = True
                        
                    if pos.get('Unterkonto') == '???':
                        desc_raw = pos.get('Beschreibung', '')
                        desc_norm = clean_description_for_dedup(desc_raw)
                        liefer_id = pos.get('Liefer ID', '')
                        kunden_id = pos.get('Kunden ID', '')
                        rechnung_kontext = pos.get('_rechnung_kontext', '')
                        
                        is_er = pos.get('Aktiv/Passiv', 'Passiva') == 'Passiva'
                        target_dict = unique_unknowns_er if is_er else unique_unknowns_ar
                        
                        # Find matching key using fuzzy logic
                        matched_key = None
                        for existing_key in target_dict.keys():
                            e_liefer, e_desc, e_kunden, e_kontext = existing_key
                            if e_liefer == liefer_id and e_kunden == kunden_id and e_kontext == rechnung_kontext:
                                if is_similar_desc(e_desc, desc_norm):
                                    matched_key = existing_key
                                    break
                                    
                        if not matched_key:
                            matched_key = (liefer_id, desc_norm, kunden_id, rechnung_kontext)
                            excluded_keys = {'Typ', 'Liefer ID', 'Kunden ID', 'Menge', 'MwSt Satz', 'Dateiname', 'Unterkonto', 'Hauptkonto', 'is_pending', '_is_ai', '_rechnung_kontext'}
                            item_id = f"er_{len(unique_unknowns_er)}" if is_er else f"ar_{len(unique_unknowns_ar)}"
                            item_data = {'id': item_id, 'Desc_Norm': desc_norm}
                            if rechnung_kontext:
                                item_data['Rechnung_Kontext'] = rechnung_kontext
                            for k, v in pos.items():
                                if k not in excluded_keys and not str(k).startswith('Einzelpreis') and not str(k).startswith('Gesamtpreis'):
                                    item_data[k] = v
                            target_dict[matched_key] = {
                                'item': item_data,
                                'indices': []
                            }
                            
                        target_dict[matched_key]['indices'].append(i)
                        
                if api_key:
                    # ER Batch
                    if unique_unknowns_er:
                        items_to_send = [u['item'] for u in unique_unknowns_er.values()]
                        total_dups = sum(len(u['indices']) for u in unique_unknowns_er.values())
                        print(f"\nSende {len(items_to_send)} Eingangsrechnungs-Artikel an die KI (Dedupliziert von {total_dups} Positionen)...")
                        ai_results = Buchung_KI.ask_gemini_batch(items_to_send, api_key, nutzerdaten_dir, is_er=True)
                        for key, data in unique_unknowns_er.items():
                            unique_id = data['item']['id']
                            if unique_id in ai_results:
                                konto = ai_results[unique_id]
                                for original_i in data['indices']:
                                    alle_positionen[original_i]['Unterkonto'] = konto
                                    alle_positionen[original_i]['_is_ai'] = True
                                    
                    # AR Batch
                    if unique_unknowns_ar:
                        items_to_send = [u['item'] for u in unique_unknowns_ar.values()]
                        total_dups = sum(len(u['indices']) for u in unique_unknowns_ar.values())
                        print(f"\nSende {len(items_to_send)} Ausgangsrechnungs-Artikel an die KI (Dedupliziert von {total_dups} Positionen)...")
                        ai_results = Buchung_KI.ask_gemini_batch(items_to_send, api_key, nutzerdaten_dir, is_er=False)
                        for key, data in unique_unknowns_ar.items():
                            unique_id = data['item']['id']
                            if unique_id in ai_results:
                                konto = ai_results[unique_id]
                                for original_i in data['indices']:
                                    alle_positionen[original_i]['Unterkonto'] = konto
                                    alle_positionen[original_i]['_is_ai'] = True

                    # Post-AI: Alle Nebenpositionen, deren Hauptposition nun von der KI kontiert wurde, erben dieses Konto
                    for pos in alle_positionen:
                        if pos.get('Unterkonto') == '???':
                            ctx = pos.get('_rechnung_kontext', '')
                            rechnungs_nr = pos.get('Rechnungsnummer', '')
                            if ctx and rechnungs_nr:
                                for other in alle_positionen:
                                    if other.get('Rechnungsnummer') == rechnungs_nr and other.get('Unterkonto') != '???':
                                        clean_other = clean_description_for_dedup(other.get('Beschreibung', ''))
                                        if clean_other == ctx or not is_generic_auxiliary(clean_other):
                                            pos['Unterkonto'] = other['Unterkonto']
                                            pos['_is_ai'] = True
                                            supplier = pos.get('Lieferant', '')
                                            clean_p = clean_description_for_dedup(pos.get('Beschreibung', ''))
                                            ck = f"{supplier} | {clean_p} [KONTEXT: {ctx}]".strip().upper()
                                            db_konten_cache[ck] = {'value': other['Unterkonto'], 'confirmed': False}
                                            break

                # Generelle Konvertierung und Hauptkonto-Ableitung
                for pos in alle_positionen:
                    c = pos.get('Unterkonto')
                    if isinstance(c, str):
                        hauptkonto = c.split('_')[0]
                        pos['Hauptkonto'] = hauptkonto
                        
                        if c.isdigit():
                            pos['Unterkonto'] = int(c)
                        if hauptkonto.isdigit():
                            pos['Hauptkonto'] = int(hauptkonto)
                    else:
                        pos['Hauptkonto'] = c

                                # Heuristik: Falls Aktiv/Passiv fehlt, anhand der häufigsten VAT raten
                all_vats = []
                for pos in alle_positionen:
                    if pos.get('Liefer ID'): all_vats.append(pos['Liefer ID'])
                    if pos.get('Kunden ID'): all_vats.append(pos['Kunden ID'])
                if all_vats:
                    from collections import Counter
                    guessed_vat = Counter(all_vats).most_common(1)[0][0]
                    for pos in alle_positionen:
                        if not pos.get('Aktiv/Passiv'):
                            if pos.get('Liefer ID') == guessed_vat:
                                pos['Aktiv/Passiv'] = 'Attiva'
                            elif pos.get('Kunden ID') == guessed_vat:
                                pos['Aktiv/Passiv'] = 'Passiva'
                            else:
                                pos['Aktiv/Passiv'] = 'Passiva'
                
                print(f"\nErstelle Excel-Datei mit {len(alle_positionen)} Positionen...")
                df = pd.DataFrame(alle_positionen)
                
                # Wenn kein einziges Kennzeichen gefunden wurde, die beiden Spalten entfernen
                has_targa = any(pos.get('Kennzeichen', '') for pos in alle_positionen)
                if not has_targa:
                    if 'Kennzeichen' in df.columns:
                        df = df.drop(columns=['Kennzeichen', 'Fahrzeugtyp'])
                
                # Split dataframe by Aktiv/Passiv
                if 'Aktiv/Passiv' in df.columns:
                    df_eingang = df[df['Aktiv/Passiv'] == 'Passiva'].copy().reset_index(drop=True)
                    df_ausgang = df[df['Aktiv/Passiv'] == 'Attiva'].copy().reset_index(drop=True)
                    df_eingang = df_eingang.drop(columns=['Aktiv/Passiv'])
                    df_ausgang = df_ausgang.drop(columns=['Aktiv/Passiv'])
                else:
                    df_eingang = df.copy().reset_index(drop=True)
                    df_ausgang = pd.DataFrame(columns=df.columns)
                
                # Excel Datei generieren
                if output_dir:
                    sammlung_ordner = output_dir
                else:
                    script_dir = os.path.dirname(os.path.abspath(__file__)) if '__file__' in globals() else os.getcwd()
                    sammlung_ordner = os.path.join(script_dir, 'Excel_Sammlung')
                    
                if not os.path.exists(sammlung_ordner):
                    os.makedirs(sammlung_ordner)
                    
                excel_path = os.path.join(sammlung_ordner, 'Gesammelte_Buchungen.xlsx')
                
                # Falls the Datei schon existiert, einen eindeutigen Namen finden
                counter = 1
                while os.path.exists(excel_path):
                    excel_path = os.path.join(sammlung_ordner, f'Gesammelte_Buchungen_{counter}.xlsx')
                    counter += 1

                writer = pd.ExcelWriter(excel_path, engine='openpyxl')
                
                sheets_to_process = []
                drop_internal = [c for c in ['_is_ai', '_rechnung_kontext'] if c in df_eingang.columns]
                if not df_eingang.empty or df_ausgang.empty: # Default if both empty
                    df_eingang_export = df_eingang.drop(columns=drop_internal) if drop_internal else df_eingang
                    df_eingang_export.to_excel(writer, index=False, sheet_name='Eingangsrechnungen')
                    sheets_to_process.append(('Eingangsrechnungen', df_eingang))
                if not df_ausgang.empty:
                    drop_internal_ausgang = [c for c in ['_is_ai', '_rechnung_kontext'] if c in df_ausgang.columns]
                    df_ausgang_export = df_ausgang.drop(columns=drop_internal_ausgang) if drop_internal_ausgang else df_ausgang
                    df_ausgang_export.to_excel(writer, index=False, sheet_name='Ausgangsrechnungen')
                    sheets_to_process.append(('Ausgangsrechnungen', df_ausgang))

                from openpyxl.styles import Font
                red_font = Font(color="FF0000", bold=True)
                euro_format = '#,##0.00 €'
                percent_format = '0.00%'

                for sheet_name, df_sheet in sheets_to_process:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Automatische Spaltenbreite (Performance-optimiert: nur erste 50 Zeilen prüfen)
                    for column_cells in worksheet.columns:
                        max_length = 0
                        column_letter = column_cells[0].column_letter
                        for cell in column_cells[:50]:
                            try:
                                if cell.value:
                                    val_str = str(cell.value)
                                    length = len(val_str)
                                    if length > max_length:
                                        max_length = length
                            except Exception as e:
                                pass
                        
                        # Breite = maximale Textlänge + Puffer (ca. 1cm)
                        adjusted_width = max_length + 6 
                        if adjusted_width > 70:  # Spalten nicht unendlich groß machen
                            adjusted_width = 70
                            
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Dynamische Spaltenindizes finden (1-basiert für openpyxl)
                    col_indices = {cell.value: idx for idx, cell in enumerate(worksheet[1], start=1)}
                    
                    einzelpreis_col = next((idx for name, idx in col_indices.items() if name and str(name).startswith('Einzelpreis')), None)
                    gesamtpreis_col = next((idx for name, idx in col_indices.items() if name and str(name).startswith('Gesamtpreis')), None)
                    mwst_col = col_indices.get('MwSt (%)')
                    conto_col = col_indices.get('Unterkonto')
                    
                    # Find AI rows for this specific sheet
                    ai_rows = set()
                    if '_is_ai' in df_sheet.columns:
                        ai_rows = set(df_sheet.index[df_sheet['_is_ai'] == True].tolist())
                    
                    for row in range(2, worksheet.max_row + 1):
                        if einzelpreis_col:
                            worksheet.cell(row=row, column=einzelpreis_col).number_format = euro_format
                        if gesamtpreis_col:
                            worksheet.cell(row=row, column=gesamtpreis_col).number_format = euro_format
                        if mwst_col:
                            worksheet.cell(row=row, column=mwst_col).number_format = percent_format
                        
                        # Dataframe index is 0-based, worksheet is 1-based and row 1 is header. So row 2 is index 0.
                        if conto_col and (row - 2) in ai_rows:
                            worksheet.cell(row=row, column=conto_col).font = red_font
                
                writer.close()
                if output_dir:
                    print(f"\n[OK] Excel erfolgreich generiert: {excel_path}")
            
                print("[PROGRESS:100]")
                
                append_new_targas_to_excel(targa_file, neue_targas_set)
                
                if fehler_log:
                    log_pfad = os.path.join(sammlung_ordner, 'Fehlgeschlagen.txt')
                    with open(log_pfad, 'w', encoding='utf-8') as f:
                        f.write("Folgende Fehler traten beim Verarbeiten auf:\n\n")
                        for err in fehler_log:
                            f.write(f"- {err}\n")
                    print(f"\nAchtung: Es gab Fehler. Details siehe: {log_pfad}")
                
                if os.name == 'nt' or sys.platform == 'win32':
                    os.startfile(excel_path)
                elif sys.platform == 'darwin':
                    import subprocess
                    subprocess.run(['open', excel_path], check=True)
                else:
                    import subprocess
                    subprocess.run(['xdg-open', excel_path], check=True)
            else:
                print("\nEs wurden keine gültigen Rechnungspositionen gefunden.")
        else:
            print("Ziehe eine oder mehrere XML- oder P7M-Dateien (Drag & Drop) auf dieses Skript-Icon, um sie zu konvertieren.")
            print("Oder ziehe einen ganzen Ordner mit XML/P7M-Dateien auf das Icon.")
    except Exception as e:
        print("\nEin unerwarteter Fehler ist aufgetreten:")
        print(traceback.format_exc())
        
        # Dieser Befehl hält das Fenster ganz am Schluss offen, egal was passiert ist
        if paths is None or paths == sys.argv[1:]:
            if sys.stdout.isatty():
                input("\nDrücke Enter zum Beenden...")

if __name__ == "__main__":
    run_conversion()