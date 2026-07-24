import sys
import os

import traceback
import datetime
from openpyxl import Workbook, load_workbook

# Utils aus dem übergeordneten Ordner laden
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from utils import get_text, safe_float, read_xml_or_p7m
import defusedxml.ElementTree as ET

import Analyse_Config
import Analyse_KI

def ensure_dashboard_template(base_dir):
    """Erstellt ein rudimentäres Dashboard Template, falls keines existiert."""
    template_path = os.path.join(base_dir, "Systemdaten", "Dashboard_Template.xlsx")
    if not os.path.exists(template_path):
        os.makedirs(os.path.dirname(template_path), exist_ok=True)
        wb = Workbook()
        
        # Dashboard Sheet
        ws_dash = wb.active
        ws_dash.title = "Dashboard"
        ws_dash['A1'] = "Sektorenanalyse Dashboard"
        ws_dash['A3'] = "Hier kannst du später Pivot-Tabellen und Diagramme einfügen, die auf den 'Rohdaten' basieren."
        
        # Rohdaten Sheet
        ws_raw = wb.create_sheet("Rohdaten")
        ws_raw['A1'] = "Rechnung"
        
        wb.save(template_path)
        print("Standard Dashboard_Template.xlsx erstellt.")
    return template_path

from sdi_parser import parse_sdi_xml

def parse_invoices(folder_path):
    alle_positionen = []
    fehler_log = []
    
    for root_dir, _, files in os.walk(folder_path):
        for filename in files:
            if not (filename.lower().endswith('.xml') or filename.lower().endswith('.p7m')):
                continue
                
            xml_path = os.path.join(root_dir, filename)
            
            # sdi_parser benötigt targa_dict und neue_targas_set (können hier leer/None sein, da für Analyse nicht zwingend)
            parsed_items = parse_sdi_xml(xml_path, targa_dict=None, neue_targas_set=None, fehler_log=fehler_log, shorten_description=False)
            
            for item in parsed_items:
                alle_positionen.append({
                    'id': str(len(alle_positionen)), # Eindeutige ID für KI
                    'Datei': filename,
                    'Datum': item['Datum'],
                    'Nummer': item['Rechnungsnummer'],
                    'Lieferant': item['Lieferant'],
                    'desc': item['Beschreibung_Full'], # In der Analyse nutzen wir oft die volle Beschreibung
                    'Menge': item['Menge'],
                    'Gesamtpreis': item['Gesamtpreis_Roh']
                })
                
    return alle_positionen

def run_analyse(rechnungs_ordner, kunden_name, base_dir, nutzerdaten_dir):
    print(f"\n--- Starte Sektorenanalyse für: {kunden_name} ---")
    
    # 1. Setup prüfen / laden
    Analyse_Config.ensure_setup_file(nutzerdaten_dir)
    kategorien = Analyse_Config.load_setup(nutzerdaten_dir)
    if not kategorien:
        print("Fehler: Keine Kategorien in Analyse_Setup.xlsx definiert!")
        return
        
    system_instruction = Analyse_Config.build_system_instruction(kategorien)
    
    # 2. Rechnungen einlesen
    print(f"Lese Rechnungen aus {rechnungs_ordner}...")
    positionen = parse_invoices(rechnungs_ordner)
    if not positionen:
        print("Keine gültigen Rechnungspositionen gefunden.")
        return
        
    print(f"{len(positionen)} Positionen gefunden.")
    
    # 3. KI Analyse
    api_key = Analyse_KI.get_api_key(base_dir)
    ki_results = Analyse_KI.analyze_items_with_ai(positionen, api_key, nutzerdaten_dir, system_instruction)
    
    # 4. Daten zusammenführen
    kat_names = [k['name'] for k in kategorien]
    
    for pos in positionen:
        pos_id = pos['id']
        ki_data = ki_results.get(pos_id, {})
        for kat in kat_names:
            pos[kat] = ki_data.get(kat, "Unbekannt")
            
    # 5. Excel schreiben (Template Injection)
    template_path = ensure_dashboard_template(base_dir)
    ausgabe_ordner = os.path.join(base_dir, "Kunden", kunden_name, "Analysen")
    os.makedirs(ausgabe_ordner, exist_ok=True)
    
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    ausgabe_datei = os.path.join(ausgabe_ordner, f"Sektorenstudie_{timestamp}.xlsx")
    
    try:
        import shutil
        shutil.copy2(template_path, ausgabe_datei)
        
        wb = load_workbook(ausgabe_datei)
        if "Rohdaten" not in wb.sheetnames:
            wb.create_sheet("Rohdaten")
        ws_raw = wb["Rohdaten"]
        
        # Existierende Inhalte in Rohdaten löschen (falls Vorlage gefüllt war)
        ws_raw.delete_rows(1, ws_raw.max_row)
        
        # Header schreiben
        headers = ["Datei", "Datum", "Nummer", "Lieferant", "Beschreibung", "Menge", "Gesamtpreis"] + kat_names
        ws_raw.append(headers)
        
        from openpyxl.styles import Font
        for cell in ws_raw[1]:
            cell.font = Font(bold=True)
            
        # Daten schreiben
        for pos in positionen:
            row_data = [
                pos['Datei'], pos['Datum'], pos['Nummer'], pos['Lieferant'], pos['desc'], 
                pos['Menge'], pos['Gesamtpreis']
            ]
            for kat in kat_names:
                row_data.append(pos[kat])
            ws_raw.append(row_data)
            
        # Formatierung (Preise und Mengen als Zahlen)
        euro_format = '#,##0.00 €'
        for row in range(2, ws_raw.max_row + 1):
            ws_raw.cell(row=row, column=6).number_format = '0.00' # Menge
            ws_raw.cell(row=row, column=7).number_format = euro_format # Preis
            
        wb.save(ausgabe_datei)
        print("\n✅ Analyse erfolgreich abgeschlossen!")
        print(f"Ergebnis gespeichert unter:\n{ausgabe_datei}")
        
        # Datei öffnen
        if os.name == 'nt' or sys.platform == 'win32':
            os.startfile(ausgabe_datei)
        elif sys.platform == 'darwin':
            import subprocess
            subprocess.run(['open', ausgabe_datei], check=True)
            
    except Exception as e:
        print(f"Fehler beim Speichern der Excel-Datei: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    if len(sys.argv) < 5:
        print("Nutzung: python Analyse_Main.py <Rechnungs_Ordner> <Kunden_Name> <Base_Dir> <Nutzerdaten_Dir>")
        sys.exit(1)
        
    run_analyse(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
