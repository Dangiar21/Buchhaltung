import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

def get_setup_path(nutzerdaten_dir):
    return os.path.join(nutzerdaten_dir, "Analyse_Setup.xlsx")

def ensure_setup_file(nutzerdaten_dir):
    """Prüft, ob die Analyse_Setup.xlsx existiert, und erstellt sie bei Bedarf."""
    setup_path = get_setup_path(nutzerdaten_dir)
    
    if not os.path.exists(setup_path):
        os.makedirs(nutzerdaten_dir, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "KI Kategorien"
        
        # Header
        ws.append(["Kategorie", "Regel für KI", "Beispiele (Optional)"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            
        # Spaltenbreiten
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 40
        
        # Beispiel eintragen
        ws.append(["Maschinenteile", "Wähle aus: Motor, Karosserie, Elektronik, Sonstiges", "Zylinderkopf -> Motor, Scheinwerfer -> Elektronik"])
        ws.append(["Qualität", "Wähle aus: Originalteil, Nachbau, Unbekannt", ""])
        
        wb.save(setup_path)
        print(f"Neue Analyse_Setup.xlsx Vorlage in {nutzerdaten_dir} erstellt.")
        return True
    return False

def load_setup(nutzerdaten_dir):
    """Lädt die Setup-Excel und gibt eine Liste von Dicts zurück."""
    setup_path = get_setup_path(nutzerdaten_dir)
    if not os.path.exists(setup_path):
        return []
        
    try:
        import polars as pl
        df = pl.read_excel(setup_path)
        kategorien = []
        for row in df.iter_rows():
            kategorie = str(row[0]).strip() if row[0] is not None else ""
            regel = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            beispiele = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            
            if kategorie and kategorie != 'nan' and kategorie != 'None':
                kategorien.append({
                    "name": kategorie,
                    "regel": regel if regel != 'nan' and regel != 'None' else "",
                    "beispiele": beispiele if beispiele != 'nan' and beispiele != 'None' else ""
                })
        return kategorien
    except Exception as e:
        print(f"Fehler beim Laden von Analyse_Setup.xlsx: {e}")
        return []

def build_system_instruction(kategorien, is_stage2=False):
    """Baut den KI-Prompt basierend auf den geladenen Kategorien."""
    if not kategorien:
        return "Du bist ein Buchhaltungs-Assistent. Keine Kategorien definiert."
        
    prompt = "Du bist ein intelligenter Datenanalyst für Sektorenstudien. Deine Aufgabe ist es, Rechnungspositionen zu kategorisieren.\n"
    prompt += "WICHTIG: Erfinde KEINE eigenen Kategorien. Nutze 'Sonstiges' oder 'Unbekannt', wenn etwas nicht zuordnenbar ist.\n\n"
    prompt += "Hier sind die Kategorien und Regeln:\n\n"
    
    kategorien_dict = {}
    for kat in kategorien:
        prompt += f"--- {kat['name']} ---\n"
        if kat['regel']:
            prompt += f"Regel: {kat['regel']}\n"
        if kat['beispiele']:
            prompt += f"Beispiele: {kat['beispiele']}\n"
        prompt += "\n"
        kategorien_dict[kat['name']] = "Dein_Ergebnis_Hier"
        
    prompt += "Du erhältst eine Liste von Artikeln im Format: [ID] Lieferant | Beschreibung\n"
    prompt += "Antworte AUSSCHLIESSLICH mit einem validen JSON-Objekt, das eine Liste unter dem Schlüssel 'ergebnisse' enthält.\n"
    
    import json
    if not is_stage2:
        prompt += "Jedes Objekt in der Liste MUSS exakt diese 4 Schlüssel haben: 'id' (als String), 'gedankengang', 'konfidenz', 'konto'.\n"
        prompt += "'gedankengang' ist ein Satz, 'konfidenz' ist eine Zahl 1-10.\n"
        prompt += "'konto' (wir nennen es intern kategorien_werte) ist ein JSON-Objekt mit den Kategorien als Schlüssel.\n"
        
        beispiel_antwort = {
            "ergebnisse": [
                {
                    "id": "0",
                    "gedankengang": "Erklaerung hier",
                    "konfidenz": 9,
                    "konto": kategorien_dict
                }
            ]
        }
    else:
        prompt += "Jedes Objekt in der Liste MUSS diese Schlüssel haben: 'id' (als String) und 'konto'.\n"
        prompt += "Der Schlüssel 'konto' ist ein Objekt mit den Kategorien als Schlüssel.\n"
        prompt += "Du MUSST dich zwingend für jede Kategorie entscheiden. Lass den Wert NIEMALS leer, auch wenn du unsicher bist.\n"
        beispiel_antwort = {
            "ergebnisse": [
                {
                    "id": "0",
                    "konto": kategorien_dict
                }
            ]
        }
        
    prompt += f"\nBeispiel-Antwort:\n{json.dumps(beispiel_antwort, indent=2)}"
    
    return prompt
