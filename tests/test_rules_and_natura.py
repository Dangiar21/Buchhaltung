import unittest
import os
import tempfile
import shutil
import openpyxl
import pandas as pd
import sys

# Ensure modules can be found
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(os.path.join(base_dir, "Programme"))
sys.path.append(os.path.join(base_dir, "Programme", "Buchungen erstellen"))

import Buchung_Regeln
from sdi_parser import parse_sdi_xml

class TestRulesAndNatura(unittest.TestCase):

    def setUp(self):
        self.test_dir = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def test_rule_matching_priorities(self):
        """Testet die 6-stufige Prioritätskette und die Kombi-Logik (Lieferant UND Stichwort)."""
        rules = {
            "client_kombi": [
                {"lieferant": "amazon", "suchbegriff": "monitor", "konto": "0650"},
                {"lieferant": "amazon", "suchbegriff": "papier", "konto": "4000"},
            ],
            "client_lieferant": [
                {"lieferant": "eni", "suchbegriff": "", "konto": "4300"},
                {"lieferant": "amazon", "suchbegriff": "", "konto": "4999"}, # Allgemeines Amazon-Konto
            ],
            "client_stichwort": [
                {"lieferant": "", "suchbegriff": "diesel", "konto": "4300"},
            ],
            "global_kombi": [
                {"lieferant": "telekom", "suchbegriff": "router", "konto": "0651"},
            ],
            "global_stichwort": [
                {"lieferant": "", "suchbegriff": "miete", "konto": "7100"},
            ],
            "global_lieferant": [
                {"lieferant": "telekom", "suchbegriff": "", "konto": "7300"},
            ]
        }

        # 1. Client Kombi (Amazon + Monitor) -> 0650
        konto, is_pending = Buchung_Regeln.assign_account(
            "DELL MONITOR 27 ZOLL", "Dell Monitor 27 Zoll", "Amazon EU S.a.r.l.", "LU123456", "KUNDE1", rules
        )
        self.assertEqual(konto, "0650")
        self.assertFalse(is_pending)

        # 2. Client Kombi (Amazon + Papier) -> 4000
        konto, _ = Buchung_Regeln.assign_account(
            "DRUCKER PAPIER A4", "Drucker Papier A4", "Amazon EU S.a.r.l.", "LU123456", "KUNDE1", rules
        )
        self.assertEqual(konto, "4000")

        # 3. Client Nur-Lieferant (Amazon mit unbekanntem Produkt -> Fallback auf 4999)
        konto, _ = Buchung_Regeln.assign_account(
            "KAFFEEBOHNEN", "Kaffeebohnen", "Amazon EU S.a.r.l.", "LU123456", "KUNDE1", rules
        )
        self.assertEqual(konto, "4999")

        # 4. Client Nur-Lieferant (Eni ohne Stichwort) -> 4300
        konto, _ = Buchung_Regeln.assign_account(
            "TANKUNG SCHALF", "Tankung Schalf", "Eni Station Meran", "", "KUNDE1", rules
        )
        self.assertEqual(konto, "4300")

        # 5. Client Nur-Stichwort (Diesel von beliebigem Lieferanten) -> 4300
        konto, _ = Buchung_Regeln.assign_account(
            "DIESEL EXTRA", "Diesel Extra", "Freie Tankstelle", "", "KUNDE1", rules
        )
        self.assertEqual(konto, "4300")

        # 6. Global Kombi (Telekom + Router) -> 0651
        konto, _ = Buchung_Regeln.assign_account(
            "SPEEDPORT ROUTER VDSL", "Speedport Router VDSL", "Telekom Deutschland", "", "KUNDE1", rules
        )
        self.assertEqual(konto, "0651")

        # 7. Global Nur-Stichwort (Miete) -> 7100
        konto, _ = Buchung_Regeln.assign_account(
            "BUEROMIETE SEPTEMBER", "Bueromiete September", "Immobilien GmbH", "", "KUNDE1", rules
        )
        self.assertEqual(konto, "7100")

        # 8. Global Nur-Lieferant (Telekom allgemein ohne Router) -> 7300
        konto, _ = Buchung_Regeln.assign_account(
            "FESTNETZ TELEFONIE", "Festnetz Telefonie", "Telekom Deutschland", "", "KUNDE1", rules
        )
        self.assertEqual(konto, "7300")

        # 9. Kein Match -> ???
        konto, _ = Buchung_Regeln.assign_account(
            "UNBEKANNTES PRODUKT", "Unbekanntes Produkt", "Unbekannter Shop", "", "KUNDE1", rules
        )
        self.assertEqual(konto, "???")

    def test_ensure_rule_file_migration(self):
        """Testet, dass alte Excel-Dateien mit Lieferanten- und Stichwort-Reiter automatisch migriert werden."""
        old_file = os.path.join(self.test_dir, "OldRules.xlsx")
        wb = openpyxl.Workbook()
        ws_lief = wb.active
        ws_lief.title = "Lieferanten-Regeln"
        ws_lief.append(["Lieferant (Name oder MwSt-Nr)", "Konto"])
        ws_lief.append(["Eni SpA", "4300"])
        ws_lief.append(["IT01234567890", "7000"])

        ws_stich = wb.create_sheet("Stichwort-Regeln")
        ws_stich.append(["Stichwort in Beschreibung", "Konto"])
        ws_stich.append(["Diesel", "4300"])
        ws_stich.append(["Porto", "7400"])

        ws_old_ki = wb.create_sheet("KI-Zuweisungen")
        ws_old_ki.append(["Dummy", "Daten"])

        wb.save(old_file)

        # Migration ausführen
        Buchung_Regeln.ensure_rule_file(old_file)

        # Überprüfen
        wb_migrated = openpyxl.load_workbook(old_file)
        self.assertEqual(wb_migrated.sheetnames, ["Regeln"])
        ws_r = wb_migrated["Regeln"]

        rows = list(ws_r.iter_rows(values_only=True))
        # Header
        self.assertEqual(rows[0], ("Lieferant (Name oder MwSt-Nr)", "Stichwort in Beschreibung", "Konto"))

        # Prüfe migrierte Daten
        data_rows = [(r[0] or "", r[1] or "", r[2] or "") for r in rows[1:]]
        self.assertIn(("Eni SpA", "", "4300"), data_rows)
        self.assertIn(("IT01234567890", "", "7000"), data_rows)
        self.assertIn(("", "Diesel", "4300"), data_rows)
        self.assertIn(("", "Porto", "7400"), data_rows)

        # Parsen testen
        parsed = Buchung_Regeln.parse_excel_rules(old_file)
        self.assertEqual(len(parsed), 4)

    def test_xml_natura_extraction(self):
        """Testet das Auslesen des Natura-Tags in DettaglioLinee sowie Fallback über DatiRiepilogo."""
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
<p:FatturaElettronica xmlns:p="http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2" versione="FPR12">
    <FatturaElettronicaHeader>
        <CedentePrestatore>
            <DatiAnagrafici>
                <IdFiscaleIVA><IdPaese>IT</IdPaese><IdCodice>01234567890</IdCodice></IdFiscaleIVA>
                <Anagrafica><Denominazione>Test Lieferant GmbH</Denominazione></Anagrafica>
            </DatiAnagrafici>
        </CedentePrestatore>
        <CessionarioCommittente>
            <DatiAnagrafici>
                <IdFiscaleIVA><IdPaese>IT</IdPaese><IdCodice>09876543210</IdCodice></IdFiscaleIVA>
                <Anagrafica><Denominazione>Test Kunde SpA</Denominazione></Anagrafica>
            </DatiAnagrafici>
        </CessionarioCommittente>
    </FatturaElettronicaHeader>
    <FatturaElettronicaBody>
        <DatiGenerali>
            <DatiGeneraliDocumento>
                <TipoDocumento>TD01</TipoDocumento>
                <Divisa>EUR</Divisa>
                <Data>2026-08-15</Data>
                <Numero>123/2026</Numero>
                <DatiBollo>
                    <BolloVirtuale>SI</BolloVirtuale>
                    <ImportoBollo>2.00</ImportoBollo>
                </DatiBollo>
            </DatiGeneraliDocumento>
        </DatiGenerali>
        <DatiBeniServizi>
            <DettaglioLinee>
                <NumeroLinea>1</NumeroLinea>
                <Descrizione>Exportware nach Schweiz</Descrizione>
                <Quantita>1.00</Quantita>
                <PrezzoUnitario>1000.00</PrezzoUnitario>
                <PrezzoTotale>1000.00</PrezzoTotale>
                <AliquotaIVA>0.00</AliquotaIVA>
                <Natura>N3.1</Natura>
            </DettaglioLinee>
            <DettaglioLinee>
                <NumeroLinea>2</NumeroLinea>
                <Descrizione>Reverse Charge Bauleistung</Descrizione>
                <Quantita>1.00</Quantita>
                <PrezzoUnitario>500.00</PrezzoUnitario>
                <PrezzoTotale>500.00</PrezzoTotale>
                <AliquotaIVA>0.00</AliquotaIVA>
            </DettaglioLinee>
            <DettaglioLinee>
                <NumeroLinea>3</NumeroLinea>
                <Descrizione>Normale Ware 22%</Descrizione>
                <Quantita>1.00</Quantita>
                <PrezzoUnitario>200.00</PrezzoUnitario>
                <PrezzoTotale>200.00</PrezzoTotale>
                <AliquotaIVA>22.00</AliquotaIVA>
            </DettaglioLinee>
            <DatiRiepilogo>
                <AliquotaIVA>0.00</AliquotaIVA>
                <Natura>N6.7</Natura>
                <ImponibileImporto>500.00</ImponibileImporto>
                <Imposta>0.00</Imposta>
            </DatiRiepilogo>
            <DatiRiepilogo>
                <AliquotaIVA>22.00</AliquotaIVA>
                <ImponibileImporto>200.00</ImponibileImporto>
                <Imposta>44.00</Imposta>
            </DatiRiepilogo>
        </DatiBeniServizi>
    </FatturaElettronicaBody>
</p:FatturaElettronica>
"""
        xml_path = os.path.join(self.test_dir, "test_natura.xml")
        with open(xml_path, "w", encoding="utf-8") as fp:
            fp.write(xml_content)

        items = parse_sdi_xml(xml_path, {}, set(), [])
        self.assertEqual(len(items), 4) # 3 Zeilen + 1 Marca da Bollo

        # Zeile 1: Natura direkt aus DettaglioLinee ("N3.1")
        self.assertEqual(items[0]['Beschreibung'], "Exportware nach Schweiz")
        self.assertEqual(items[0]['MwSt'], 0.0)
        self.assertEqual(items[0]['Natura'], "N3.1")

        # Zeile 2: Natura per Fallback aus DatiRiepilogo ("N6.7")
        self.assertEqual(items[1]['Beschreibung'], "Reverse Charge Bauleistung")
        self.assertEqual(items[1]['MwSt'], 0.0)
        self.assertEqual(items[1]['Natura'], "N6.7")

        # Zeile 3: Normale MwSt 22% -> Natura leer
        self.assertEqual(items[2]['Beschreibung'], "Normale Ware 22%")
        self.assertEqual(items[2]['MwSt'], 0.22)
        self.assertEqual(items[2]['Natura'], "")

        # Zeile 4: Marca da Bollo (DatiBollo, MwSt 0%, übernimmt Fallback oder leer)
        self.assertEqual(items[3]['Beschreibung'], "Marca da Bollo")
        self.assertEqual(items[3]['MwSt'], 0.0)

if __name__ == '__main__':
    unittest.main()
