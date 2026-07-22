import openpyxl
import os
import sqlite3

global_path = 'c:/Users/Dangi/Desktop/Buchhaltung/Systemdaten/Globale_KontenRegeln.xlsx'

if os.path.exists(global_path):
    wb = openpyxl.load_workbook(global_path)
    if "KI-Zuweisungen" in wb.sheetnames:
        ws = wb["KI-Zuweisungen"]
        rows_to_delete = []
        for row in range(2, ws.max_row + 1):
            konto = ws.cell(row=row, column=6).value
            if str(konto).strip() == '5000000':
                rows_to_delete.append(row)
        
        # Delete from bottom to top
        for row in reversed(rows_to_delete):
            ws.delete_rows(row)
            
        wb.save(global_path)
        print(f"Deleted {len(rows_to_delete)} rows from KI-Zuweisungen in Excel.")

db_path = 'c:/Users/Dangi/Desktop/Buchhaltung/Systemdaten/buchhaltung.db'
if os.path.exists(db_path):
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("DELETE FROM kontenregeln WHERE konto='5000000'")
    conn.commit()
    print(f"Deleted {cur.rowcount} rows from kontenregeln in SQLite.")
    
    cur.execute("DELETE FROM cache_konten WHERE konto='5000000'")
    conn.commit()
    print(f"Deleted {cur.rowcount} rows from cache_konten in SQLite.")
    conn.close()
